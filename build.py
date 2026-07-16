# -*- coding: utf-8 -*-
"""
build.py — единая точка пересборки базы авто остановов ГПА.

Использование:
    python3 build.py

Логика:
    1. Читает единственный источник данных accidents.json
    2. Генерирует xlsx (5 листов, часть — формулами)
    3. Генерирует html (данные встроены в JS-массив, но собраны из accidents.json)

Чтобы добавить новую аварию — редактируется ТОЛЬКО accidents.json,
затем запускается этот скрипт. Ничего больше редактировать вручную не нужно.
"""
import json
import html as html_lib
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / 'accidents.json'
DEFECTS_FILE = BASE_DIR / 'defects.json'
XLSX_OUT = BASE_DIR / 'База_аварийных_остановов_ГПА.xlsx'
HTML_OUT = BASE_DIR / 'Поиск_по_базе_аварий.html'

FONT = 'Arial'
GREY = 'FFF7FAFC'
BORDER = Border(*(Side(style='thin', color='FFD0D0D0'),) * 4)


import base64


def pdf_to_data_uri(filename):
    """Кодирует PDF-файл (лежащий рядом с build.py) в data-URI, чтобы ссылка в html
    работала без внешнего файла на диске."""
    path = BASE_DIR / filename
    if not path.exists():
        return None
    data = path.read_bytes()
    b64 = base64.b64encode(data).decode('ascii')
    return f'data:application/pdf;base64,{b64}'


WEB_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta http-equiv="X-UA-Compatible" content="IE=edge">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Поиск по базе авто остановов ГПА</title>
<style>
  :root{{
    --navy:#1A365D; --navy-dark:#0F2440; --surface:#F7FAFC; --card:#FFFFFF;
    --border:#DDE3EA; --text:#1F2933; --text-muted:#5C6B7A;
    --teal:#0F6E56; --teal-bg:#E1F5EE; --coral:#993C1D; --coral-bg:#FAECE7;
    --amber:#854F0B; --amber-bg:#FAEEDA;
    --mono:"SF Mono","Consolas","Roboto Mono",monospace;
    --sans:"Segoe UI",Inter,system-ui,-apple-system,Arial,sans-serif;
  }}
  *{{box-sizing:border-box;}}
  body{{margin:0;background:var(--surface);color:var(--text);font-family:var(--sans);line-height:1.5;}}
  header{{background:linear-gradient(180deg,var(--navy),var(--navy-dark));color:#fff;padding:28px 24px 22px;}}
  .header-row{{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;}}
  header h1{{margin:0 0 4px;font-size:20px;font-weight:600;}}
  header p{{margin:0;font-size:13px;color:#C9D6E5;}}
  .lang-btn{{flex:0 0 auto;padding:8px 16px;font-size:13px;font-weight:700;border:1px solid rgba(255,255,255,0.35);border-radius:8px;background:rgba(255,255,255,0.08);color:#fff;cursor:pointer;}}
  .lang-btn:hover{{background:rgba(255,255,255,0.18);}}
  .lang-btn-toolbar{{background:var(--surface);color:var(--navy);border:1px solid var(--border);}}
  .lang-btn-toolbar:hover{{background:#EDF1F5;}}
  .container{{max-width:960px;margin:0 auto;padding:20px 20px 60px;}}
  .toolbar{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px;margin:-18px 0 20px;box-shadow:0 2px 10px rgba(15,36,64,0.08);}}
  .search-row{{display:flex;gap:10px;margin-bottom:12px;}}
  .search-row input{{flex:1;padding:11px 14px;font-size:14px;border:1px solid var(--border);border-radius:8px;outline:none;font-family:var(--sans);}}
  .search-row input:focus{{border-color:var(--navy);box-shadow:0 0 0 3px rgba(26,54,93,0.12);}}
  .filters{{display:flex;gap:10px;flex-wrap:wrap;}}
  .filters select{{padding:8px 10px;font-size:13px;border:1px solid var(--border);border-radius:7px;background:#fff;color:var(--text);font-family:var(--sans);}}
  .filters button{{padding:8px 14px;font-size:13px;border:1px solid var(--border);border-radius:7px;background:var(--surface);color:var(--text-muted);cursor:pointer;}}
  .filters button:hover{{background:#EDF1F5;}}
  .meta-row{{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:12px;font-size:13px;color:var(--text-muted);}}
  .meta-row b{{color:var(--text);}}
  .card{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px 18px;margin-bottom:12px;}}
  .card-top{{display:flex;gap:8px;align-items:center;margin-bottom:8px;flex-wrap:wrap;}}
  .tag{{font-family:var(--mono);font-size:11.5px;padding:3px 8px;border-radius:5px;font-weight:600;letter-spacing:0.2px;}}
  .tag.act{{background:#EDF1F5;color:var(--navy);}}
  .tag.gpa{{background:var(--teal-bg);color:var(--teal);}}
  .tag.cat{{background:var(--amber-bg);color:var(--amber);}}
  .tag.type{{background:var(--coral-bg);color:var(--coral);}}
  .card h3{{margin:0 0 6px;font-size:15.5px;font-weight:600;color:var(--text);}}
  .card .cause{{font-size:13.5px;color:var(--text-muted);margin:0 0 8px;}}
  mark{{background:#FFE9A8;color:#412402;padding:0 2px;border-radius:2px;}}
  .details{{display:none;font-size:13px;margin-top:8px;padding-top:10px;border-top:1px dashed var(--border);}}
  .details.open{{display:block;}}
  .details p{{margin:0 0 8px;}}
  .details b{{color:var(--text);}}
  .toggle-btn{{font-size:12.5px;color:var(--navy);background:none;border:none;cursor:pointer;padding:0;font-weight:600;text-decoration:underline;}}
  .empty{{text-align:center;color:var(--text-muted);padding:50px 10px;font-size:14px;}}
  .empty div{{font-size:28px;margin-bottom:8px;}}
  .src-note{{font-size:11px;color:#9AA5B1;text-align:center;margin-top:24px;}}
  a.doc-link{{color:var(--navy);font-weight:600;text-decoration:underline;}}
  .loading{{text-align:center;color:var(--text-muted);padding:60px 10px;}}
  .file-btn{{font-size:13px;font-weight:600;color:var(--navy);background:var(--surface);border:1px solid var(--border);border-radius:7px;padding:8px 14px;cursor:pointer;display:inline-block;}}
  .file-btn:hover{{background:#EDF1F5;}}
  #fileStatus{{font-size:12.5px;color:var(--text-muted);flex-basis:100%;margin-top:8px;}}
  #fileStatus.error{{color:var(--coral);}}
  .tag.match{{background:#DCFCE7;color:#166534;}}
  .equip-tags{{margin:0 0 8px;}}
  .tag.equip{{font-family:var(--mono);font-size:11px;padding:2px 7px;border-radius:5px;font-weight:600;background:#EDE9FE;color:#5B21B6;margin:0 5px 4px 0;display:inline-block;}}
  .tabs{{display:flex;gap:8px;margin:0 0 14px;}}
  .tab-btn{{flex:1;padding:11px 14px;font-size:13.5px;font-weight:600;border:1px solid var(--border);border-radius:9px;background:var(--card);color:var(--text-muted);cursor:pointer;box-shadow:0 2px 10px rgba(15,36,64,0.06);}}
  .tab-btn.active{{background:var(--navy);color:#fff;border-color:var(--navy);}}
  .tab-btn:hover:not(.active){{background:#EDF1F5;}}
  .dash-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:12px;margin-bottom:16px;}}
  .stat-card{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:14px 16px;}}
  .stat-card .num{{font-size:24px;font-weight:700;color:var(--navy);line-height:1.2;}}
  .stat-card .lbl{{font-size:12px;color:var(--text-muted);margin-top:4px;}}
  .chart-card{{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px 18px;margin-bottom:14px;}}
  .chart-card h4{{margin:0 0 14px;font-size:14px;font-weight:600;color:var(--text);}}
  .bar-row{{display:flex;align-items:center;gap:10px;margin-bottom:8px;font-size:12.5px;}}
  .bar-label{{flex:0 0 210px;color:var(--text-muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
  .bar-track{{flex:1;background:var(--surface);border-radius:5px;height:17px;position:relative;overflow:hidden;}}
  .bar-fill{{height:100%;border-radius:5px;min-width:2px;}}
  .bar-val{{flex:0 0 28px;text-align:right;font-weight:600;color:var(--text);}}
  .yearbars{{display:flex;align-items:flex-end;gap:10px;height:150px;}}
  .yearbar-col{{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:flex-end;height:100%;}}
  .yearbar{{width:100%;max-width:38px;background:var(--teal);border-radius:4px 4px 0 0;transition:height .2s;}}
  .yearbar-lbl{{font-size:11px;color:var(--text-muted);margin-top:5px;}}
  .yearbar-val{{font-size:11.5px;font-weight:600;color:var(--text);margin-bottom:3px;}}
  .rec-block{{background:var(--teal-bg);border:1px solid var(--teal);border-radius:10px;padding:16px 18px;margin-bottom:14px;}}
  .rec-title{{font-size:14px;font-weight:700;color:var(--teal);margin:0 0 10px;}}
  .rec-line{{font-size:13.5px;margin:0 0 8px;color:var(--text);}}
  .rec-line:last-child{{margin-bottom:0;}}
  .rec-line b{{color:var(--navy);}}
  .rec-note{{font-size:12.5px;margin:8px 0 0;padding-top:8px;border-top:1px dashed rgba(15,110,86,0.35);color:var(--amber);font-weight:600;}}
  .reg-row{{display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:1px solid var(--border);font-size:12.5px;}}
  .reg-row:last-child{{border-bottom:none;}}
  .reg-cat{{flex:1;color:var(--text);}}
  .reg-count{{flex:0 0 auto;background:var(--coral-bg);color:var(--coral);font-weight:700;padding:2px 9px;border-radius:12px;font-size:11.5px;white-space:nowrap;}}
  .reg-dates{{flex:0 0 150px;text-align:right;color:var(--text-muted);font-size:11.5px;}}
  .info-callout{{background:var(--amber-bg);border:1px solid var(--amber);border-radius:10px;padding:14px 16px;font-size:13px;color:var(--text);margin-bottom:14px;line-height:1.55;}}
  .info-callout b{{color:var(--amber);}}
  @media (max-width:600px){{.search-row{{flex-direction:column;}} .bar-label{{flex-basis:120px;}} .reg-dates{{flex-basis:110px;}}}}
</style>
</head>
<body>
<header>
  <div class="header-row">
    <div>
      <h1 id="pageTitle">Поиск по базе авто остановов ГПА</h1>
      <p id="pageSubtitle">Умная система анализа авто остановов и помощи инженерам &middot; КС-1 &laquo;Алимтау&raquo;</p>
    </div>
    <button id="langToggle" type="button" class="lang-btn">EN</button>
  </div>
</header>
<div class="container">
  <div class="tabs">
    <button class="tab-btn active" id="tabSearchBtn" type="button">&#128269; Поиск по остановам</button>
    <button class="tab-btn" id="tabDashBtn" type="button">&#128202; Dashboard</button>
  </div>
  <div id="searchView">
  <div class="toolbar">
    <div class="search-row">
      <input id="q" type="text" placeholder="Например: свечной кран, потеря пламени, RB6-2, ГПА№2..." autocomplete="off" disabled>
    </div>
    <div class="filters">
      <select id="fGpa" disabled><option value="" id="fGpaAll">Все ГПА</option></select>
      <select id="fCat" disabled><option value="" id="fCatAll">Все категории</option></select>
      <select id="fYear" disabled><option value="" id="fYearAll">Все годы</option></select>
      <button id="reset" type="button" disabled>Сбросить</button>
      <button id="langToggle2" type="button" class="lang-btn lang-btn-toolbar">EN</button>
      <label class="file-btn" for="reportFile" id="uploadLabel">&#128206; Загрузить файл донесения (PDF, Word, TXT, фото)</label>
      <input id="reportFile" type="file" accept=".pdf,.docx,.txt,.png,.jpg,.jpeg" style="display:none;">
      <span id="fileStatus"></span>
    </div>
  </div>
  <div class="meta-row">
    <span class="left" id="count"></span>
    <span class="right" id="sortLabel">Сортировка по релевантности</span>
  </div>
  <div id="recBlock"></div>
  <div id="results"><div class="loading">Загрузка данных...</div></div>
  <div class="src-note" id="srcNote">Данные загружаются из accidents.json + defects.json &middot; акты — (лежат в корне репозитория)</div>
  </div>
  <div id="dashboardView" style="display:none;">
    <div id="dashContent"><div class="loading">Загрузка данных...</div></div>
  </div>
</div>
<script>
var incidents = [];
var LANG = 'ru';
var BUILD_STAMP = '';
function FS(ru, en) {{ return LANG==='en' ? en : ru; }}
var CAT_EN = {{
  'Внешняя/неподтвержденная причина':'External / unconfirmed cause',
  'Ложное срабатывание КИП':'False instrumentation trip',
  'Ложное срабатывание КИП / связь':'False instrumentation trip / communication',
  'Настройка/калибровка камеры сгорания (DLE)':'Combustor tuning / calibration (DLE)',
  'Отказ оборудования КИПиА':'Instrumentation & control equipment failure',
  'Техническая неисправность (вспомогательные системы)':'Technical fault (auxiliary systems)',
  'Техническая неисправность оборудования':'Equipment technical fault',
  'Человеческий фактор':'Human factor',
  'Электросеть / переключение агрегатов':'Power grid / unit switching',
  'Электросеть / синхронизация агрегатов':'Power grid / unit synchronization',
  'Дефект (без останова)':'Defect (no shutdown)',
  'Плановая инспекция / дефект':'Planned inspection / defect'
}};
var TYPE_EN = {{'Автоматический':'Automatic','Вынужденный нормальный':'Forced normal'}};
var GPA_EN = {{'ГПА №1':'GPU #1','ГПА №2':'GPU #2','ГПА №3':'GPU #3'}};
function trCat(v) {{ return (LANG==='en' && CAT_EN[v]) ? CAT_EN[v] : v; }}
function trType(v) {{ return (LANG==='en' && TYPE_EN[v]) ? TYPE_EN[v] : v; }}
function trGpa(v) {{ return (LANG==='en' && GPA_EN[v]) ? GPA_EN[v] : v; }}
var I18N = {{
  ru: {{
    title:'Поиск по базе авто остановов ГПА',
    subtitle:'Умная система анализа авто остановов и помощи инженерам &middot; КС-1 &laquo;Алимтау&raquo;',
    tabSearch:'&#128269; Поиск по остановам', tabDash:'&#128202; Dashboard',
    searchPlaceholder:'Например: свечной кран, потеря пламени, RB6-2, ГПА№2...',
    allGpa:'Все ГПА', allCat:'Все категории', allYear:'Все годы', reset:'Сбросить',
    uploadLabel:'&#128206; Загрузить файл донесения (PDF, Word, TXT, фото)',
    sortRelevance:'Сортировка по релевантности', loading:'Загрузка данных...',
    srcNote:'Данные загружаются из accidents.json + defects.json &middot; акты — (лежат в корне репозитория)',
    circumstances:'Обстоятельства, при которых произошел останов:',
    showDetails:'Показать меры и заключение &#9662;', hideDetails:'Скрыть детали &#9652;',
    conclusion:'Заключение:', measuresTaken:'Принятые меры:',
    preventionMeasures:'Мероприятия по недопущению аналогичного автоматического останова:',
    docs:'Документы:', openInvestigationAct:'Открыть акт расследования',
    openWorkAct:'Открыть акт выполненных работ', openDefectAct:'Открыть дефектный акт',
    nothingFound:'Ничего не найдено с совпадением 25% и выше. Попробуйте другой запрос или сбросьте фильтры.',
    recTitle:'&#128161; Рекомендация на основе похожих случаев',
    recMostSimilar:'Наиболее похожий случай:', recAct:'акт', recFrom:'от', recMatch:'совпадение',
    recWhatToDo:'Что делать:', recFormalRec:'Рекомендация по итогам расследования:',
    dashStatTotal:'Всего авто остановов', dashStatPeriod:'Период наблюдения',
    dashStatTopGpa:'Чаще всего останавливался', dashStatTopCat:'Самая частая причина', times:'раз',
    dashYearChart:'Динамика авто остановов по годам', dashByGpa:'Аварии по агрегатам (ГПА)',
    dashTopCat:'Топ повторяющихся причин / категорий', dashByType:'По типу останова', dashNoData:'Нет данных',
    dashMetricsTitle:'&#128200; Метрики эффективности', dashGoal:'Цель системы:',
    dashGoalText:'снизить долю повторных авто остановов на 40% и сократить время анализа причины с 3&ndash;5 суток (ручной пролистывание архива актов) до 1 рабочего дня за счёт мгновенного поиска похожих случаев по базе.',
    dashRepeatShare:'Аварий с уже встречавшейся причиной (сейчас, без системы)',
    dashRepeatCats:'Категорий причин, повторявшихся 2+ раза',
    dashRepeatTrend:'Динамика повторяемости причин по годам (% от аварий за год)',
    dashRegistryTitle:'Реестр повторяющихся причин &mdash; кандидаты для системного устранения',
    dashRegistryEmpty:'Повторяющихся причин пока не выявлено.',
    dashRiskTitle:'&#128302; Прогноз рисков', dashRiskHow:'Как читать:',
    dashRiskHowText:'это статистическая эвристика (частота + давность последнего случая + тренд за 2 года), а не обученная ML-модель — на 30 наблюдениях полноценное машинное обучение будет либо переобучаться, либо гадать. Используйте как ориентир для приоритизации осмотров, не как замену технической диагностики.',
    dashRiskScoring:'Риск-скоринг агрегатов', riskHigh:'Высокий', riskMedium:'Средний', riskLow:'Низкий',
    dashRiskDetail:'Детализация по агрегатам', riskProbableCause:'вероятная причина:',
    riskStopAgo:'останов', riskDaysAgo:'дн. назад', riskAvgInterval:'ср. интервал', riskDays:'дн.',
    dashForecastTitle:'Прогноз срока следующего останова (по всей КС)',
    forecastAvgLabel:'Средний интервал между авариями за весь период:',
    forecastSinceLabel:'С последней аварии прошло', forecastMore:'больше', forecastLess:'меньше',
    forecastElevated:'повышена', forecastNormal:'в пределах нормы',
    noDataToShow:'Нет данных для отображения.', dashEmpty:'Нет данных для отображения.'
  }},
  en: {{
    title:'GPA Emergency Shutdown Database Search',
    subtitle:'Smart shutdown analysis and engineer support system &middot; CS-1 &laquo;Alimtau&raquo;',
    tabSearch:'&#128269; Search Shutdowns', tabDash:'&#128202; Dashboard',
    searchPlaceholder:'e.g.: ignition valve, flame loss, RB6-2, GPU#2...',
    allGpa:'All units', allCat:'All categories', allYear:'All years', reset:'Reset',
    uploadLabel:'&#128206; Upload incident report (PDF, Word, TXT, photo)',
    sortRelevance:'Sorted by relevance', loading:'Loading data...',
    srcNote:'Data loaded from accidents.json + defects.json &middot; reports are stored in the repository root',
    circumstances:'Circumstances of the shutdown:',
    showDetails:'Show measures &amp; conclusion &#9662;', hideDetails:'Hide details &#9652;',
    conclusion:'Conclusion:', measuresTaken:'Measures taken:',
    preventionMeasures:'Measures to prevent a similar automatic shutdown:',
    docs:'Documents:', openInvestigationAct:'Open investigation report',
    openWorkAct:'Open completion report', openDefectAct:'Open defect report',
    nothingFound:'Nothing found with a 25%+ match. Try a different query or reset the filters.',
    recTitle:'&#128161; Recommendation based on similar cases',
    recMostSimilar:'Most similar case:', recAct:'report', recFrom:'dated', recMatch:'match',
    recWhatToDo:'What to do:', recFormalRec:'Investigation recommendation:',
    dashStatTotal:'Total emergency shutdowns', dashStatPeriod:'Observation period',
    dashStatTopGpa:'Most frequently down', dashStatTopCat:'Most common cause', times:'times',
    dashYearChart:'Shutdowns by year', dashByGpa:'Shutdowns by unit (GPU)',
    dashTopCat:'Top recurring causes / categories', dashByType:'By shutdown type', dashNoData:'No data',
    dashMetricsTitle:'&#128200; Efficiency metrics', dashGoal:'System goal:',
    dashGoalText:'reduce the share of repeat emergency shutdowns by 40% and cut root-cause analysis time from 3&ndash;5 days (manually paging through the report archive) to 1 working day via instant search across similar past cases.',
    dashRepeatShare:'Shutdowns with an already-seen cause (current baseline, no system)',
    dashRepeatCats:'Cause categories that recurred 2+ times',
    dashRepeatTrend:'Repeat-cause trend by year (% of that year&rsquo;s shutdowns)',
    dashRegistryTitle:'Repeat-cause registry &mdash; candidates for systemic fixes',
    dashRegistryEmpty:'No recurring causes identified yet.',
    dashRiskTitle:'&#128302; Risk forecast', dashRiskHow:'How to read this:',
    dashRiskHowText:'this is a statistical heuristic (frequency + recency of last case + 2-year trend), not a trained ML model — with only 30 observations, a real machine-learning model would either overfit or just guess the majority class. Use it as a prioritization aid, not a substitute for technical diagnostics.',
    dashRiskScoring:'Unit risk scoring', riskHigh:'High', riskMedium:'Medium', riskLow:'Low',
    dashRiskDetail:'Per-unit detail', riskProbableCause:'likely cause:',
    riskStopAgo:'last shutdown', riskDaysAgo:'days ago', riskAvgInterval:'avg. interval', riskDays:'days',
    dashForecastTitle:'Forecast: next expected shutdown (whole station)',
    forecastAvgLabel:'Average interval between shutdowns over the whole period:',
    forecastSinceLabel:'Days since the last shutdown:', forecastMore:'more', forecastLess:'less',
    forecastElevated:'elevated', forecastNormal:'within normal range',
    noDataToShow:'No data to display.', dashEmpty:'No data to display.'
  }}
}};
function T(key) {{ return I18N[LANG][key]; }}

function uniqueSorted(arr) {{
  var seen = {{}}, out = [];
  for (var i = 0; i < arr.length; i++) {{ var v = arr[i]; if (!seen[v]) {{ seen[v] = true; out.push(v); }} }}
  out.sort(); return out;
}}
function arrayContains(arr, val) {{ for (var i=0;i<arr.length;i++){{ if(arr[i]===val) return true; }} return false; }}
function escapeHtml(s) {{ return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }}
function escapeRegExp(s) {{ return s.replace(/[.*+?^${{}}()|[\\]\\\\]/g,'\\\\$&'); }}

function initFilters() {{
  var allGpa=[], allCat=[], allYear=[];
  for (var i=0;i<incidents.length;i++) {{
    var inc = incidents[i];
    for (var j=0;j<inc.gpa.length;j++) {{ allGpa.push(inc.gpa[j]); }}
    allCat.push(inc.category);
    allYear.push(inc.date.slice(-4));
  }}
  fillSelect('fGpa', allGpa); fillSelect('fCat', allCat); fillSelect('fYear', allYear);
  document.getElementById('q').disabled = false;
  document.getElementById('fGpa').disabled = false;
  document.getElementById('fCat').disabled = false;
  document.getElementById('fYear').disabled = false;
  document.getElementById('reset').disabled = false;
}}
function fillSelect(id, values) {{
  var sel = document.getElementById(id);
  var uniq = uniqueSorted(values);
  for (var i=0;i<uniq.length;i++) {{
    var o = document.createElement('option');
    o.value = uniq[i]; o.textContent = uniq[i];
    sel.appendChild(o);
  }}
}}

var CYR2LAT_MAP = {{'а':'a','в':'b','е':'e','к':'k','м':'m','н':'h','о':'o','р':'p','с':'c','т':'t','у':'y','х':'x'}};
function canonicalizeTagRun(run) {{
  var out = '';
  for (var i=0;i<run.length;i++) {{ var ch = run[i]; out += CYR2LAT_MAP[ch] || ch; }}
  return out;
}}
function normalize(s) {{
  var t = (s||'').toLowerCase();
  t = t.replace(/\\b([a-zа-я]{{1,4}})[\\s-]+(\\d+)/gi,'$1$2');
  t = t.replace(/(\\d)[\\s-]+([a-zа-я]{{1,4}}\\d+)/gi,'$1$2');
  t = t.replace(/[«»"'.,()–—]/g,' ');
  t = t.replace(/-/g,'');
  t = t.replace(/[a-zа-я0-9\\/]{{3,}}/gi, function(run){{
    if (/[0-9]/.test(run) && /[a-zа-я]/i.test(run)) {{ return canonicalizeTagRun(run); }}
    return run;
  }});
  return t.replace(/^\\s+|\\s+$/g,'');
}}
var STOPWORDS = {{'и':1,'в':1,'на':1,'с':1,'по':1,'для':1,'из':1,'к':1,'от':1,'о':1,'об':1,'что':1,'это':1,'как':1,'при':1,'до':1,'за':1,'не':1,'или':1,'а':1,'но':1,'же':1,'то':1,'его':1,'её':1,'их':1,'был':1,'было':1,'были':1,'есть':1,'также':1,'более':1,'менее':1,'фио':1,'подпись':1,'дата':1,'фамилия':1,'имя':1,'отчество':1,'наименование':1,'должность':1,'инженер':1,'инженера':1,'инженеру':1,'начальник':1,'начальника':1,'служба':1,'службы':1,'службу':1,'главный':1,'главного':1,'заместитель':1,'заместителя':1,'ведущий':1,'ведущего':1,'сменный':1,'сменного':1,'шымкент':1,'алимтау':1,'утг':1,'ктг':1,'приложение':1,'attachment':1,'name':1,'signature':1,'date':1,'engineer':1,'head':1,'chief':1,'deputy':1,'service':1,'leading':1,'shift':1,'the':1,'and':1,'of':1,'to':1,'in':1,'for':1,'is':1,'are':1,'акт':1,'акта':1,'акту':1}};
function tokenize(s) {{
  var norm = normalize(s); if (norm==='') return [];
  var parts = norm.split(/\\s+/), out=[], seen={{}};
  for (var i=0;i<parts.length;i++) {{
    var t = parts[i];
    if (t==='' || t.length<2) continue;
    if (STOPWORDS[t]) continue;
    if (seen[t]) continue;
    seen[t] = true;
    out.push(t);
  }}
  return out;
}}
var COMMON_UNIT_RE = /^(гпа|гпэс|адэс|ибп|вк|ак|мг|кс)№?\d{{0,2}}$/;
var FORCED_TAGS = {{}};
var ACRONYM_STOPWORDS = {{'act':1,'the':1,'and':1,'for':1,'attachment':1,'technical':1,'investigation':1,'circumstances':1,'during':1,'which':1,'shutdown':1,'has':1,'occurred':1,'actions':1,'taken':1,'recover':1,'operating':1,'mode':1,'online':1,'damaged':1,'units':1,'details':1,'devices':1,'measures':1,'performed':1,'eliminate':1,'reasons':1,'conclusion':1,'applicability':1,'recovered':1,'further':1,'operation':1,'prevent':1,'similar':1,'auto':1,'notes':1,'name':1,'signature':1,'date':1,'engineer':1,'head':1,'chief':1,'deputy':1,'service':1,'leading':1,'shift':1,'gpa':1,'gcu':1,'report':1,'basis':1,'stop':1,'dated':1,'from':1,'about':1,'automatic':1,'forced':1,'normal':1,'emergency':1,'condition':1,'committe':1,'consisting':1,'restoration':1,'main':1,'gas':1,'pipeline':1}};
function extractUpperAcronyms(s) {{
  var matches = (s||'').match(/\\b[A-Z]{{3,8}}\\b/g) || [];
  var out = {{}};
  for (var i=0;i<matches.length;i++) {{
    var w = matches[i].toLowerCase();
    if (ACRONYM_STOPWORDS[w]) continue;
    out[w] = true;
  }}
  return out;
}}
function isTagToken(t) {{
  if (FORCED_TAGS[t]) {{ return true; }}
  if (COMMON_UNIT_RE.test(t)) {{ return false; }}
  return t.length>=3 && /[0-9]/.test(t) && /[a-zа-я]/i.test(t);
}}
var FIELD_WEIGHTS = [['cause',4],['name',2],['category',1.5],['act',1],['date',1],['gpaStr',1],['measures',1],['conclusion',0.5],['recommendation',0.5],['prevention',0.5]];
var TAG_BOOST = 9;
function scoreIncident(inc, tokens, allowWords) {{
  if (tokens.length===0) return 0.0001;
  var hasTags = false;
  for (var q=0;q<tokens.length;q++) {{ if (isTagToken(tokens[q])) {{ hasTags = true; break; }} }}
  var fields = {{name:inc.name,cause:inc.cause,category:inc.category,act:inc.act,date:inc.date,gpaStr:inc.gpa.join(' '),measures:inc.measures,conclusion:inc.conclusion,recommendation:inc.recommendation,prevention:inc.prevention_measures||''}};
  var score=0;
  var normTags = (inc.tags||[]).map(function(x){{ return normalize(x); }});
  for (var t=0;t<tokens.length;t++) {{
    var tok = tokens[t];
    var tag = isTagToken(tok);
    if (hasTags && !tag && !allowWords) continue;
    var boost = tag ? TAG_BOOST : 1;
    if (tag) {{
      for (var nt=0;nt<normTags.length;nt++) {{
        if (normTags[nt].indexOf(tok)!==-1 || tok.indexOf(normTags[nt])!==-1) {{ score += 25; break; }}
      }}
    }}
    for (var f=0;f<FIELD_WEIGHTS.length;f++) {{
      var fn=FIELD_WEIGHTS[f][0], w=FIELD_WEIGHTS[f][1];
      var text = normalize(fields[fn]);
      if (text.indexOf(tok)!==-1) {{
        var mult = boost;
        if (fn==='cause') {{ mult *= (tag ? 2 : 1.5); }}
        score += w*mult;
      }}
    }}
  }}
  return score;
}}
function highlight(text, tokens) {{
  var safe = escapeHtml(text);
  if (!tokens.length) return safe;
  for (var i=0;i<tokens.length;i++) {{
    var tok = tokens[i]; if (tok.length<2) continue;
    var re = new RegExp('('+escapeRegExp(tok)+')','ig');
    safe = safe.replace(re, '<mark>$1</mark>');
  }}
  return safe;
}}

function buildCard(inc, tokens, scoreBadge) {{
  var card = document.createElement('div');
  card.className = 'card';
  var gpaTags = '';
  for (var g=0;g<inc.gpa.length;g++) {{ gpaTags += '<span class="tag gpa">'+escapeHtml(trGpa(inc.gpa[g]))+'</span>'; }}
  var htmlStr = '';
  htmlStr += '<div class="card-top">';
  htmlStr += '<span class="tag act">'+escapeHtml(inc.act)+' &middot; '+escapeHtml(inc.date)+'</span>';
  htmlStr += gpaTags;
  htmlStr += '<span class="tag cat">'+escapeHtml(trCat(inc.category))+'</span>';
  htmlStr += '<span class="tag type">'+escapeHtml(trType(inc.type))+'</span>';
  if (scoreBadge) {{ htmlStr += '<span class="tag match">'+escapeHtml(scoreBadge)+'</span>'; }}
  htmlStr += '</div>';
  htmlStr += '<h3>'+highlight(inc.name, tokens)+'</h3>';
  htmlStr += '<p class="cause"><b>'+T('circumstances')+'</b> '+highlight(inc.cause, tokens)+'</p>';
  if (inc.tags && inc.tags.length) {{
    var tagBadges = '';
    for (var tg=0; tg<inc.tags.length; tg++) {{ tagBadges += '<span class="tag equip">'+highlight(inc.tags[tg], tokens)+'</span>'; }}
    htmlStr += '<p class="equip-tags">'+tagBadges+'</p>';
  }}
  htmlStr += '<button class="toggle-btn" type="button">'+T('showDetails')+'</button>';
  htmlStr += '<div class="details">';
  if (inc.conclusion) {{ htmlStr += '<p><b>'+T('conclusion')+'</b> '+highlight(inc.conclusion, tokens)+'</p>'; }}
  var principalMeasures = '';
  if (inc.remediation && inc.remediation.length) {{
    var pmParts = [];
    for (var w=0;w<inc.remediation.length;w++) {{
      var wk = inc.remediation[w];
      pmParts.push((wk.work_description||'')+' — '+FS('Акт выполненных работ','Completion report')+' '+(wk.work_act||'?')+' '+FS('от','dated')+' '+(wk.work_date||'?')+' ('+(wk.status||FS('выполнено','completed'))+')');
    }}
    principalMeasures = pmParts.join('<br>');
  }} else {{
    principalMeasures = inc.measures || '—';
  }}
  htmlStr += '<p><b>'+T('measuresTaken')+'</b> '+highlight(principalMeasures, tokens)+'</p>';
  if (inc.prevention_measures) {{
    htmlStr += '<p><b>'+T('preventionMeasures')+'</b> '+highlight(inc.prevention_measures, tokens)+'</p>';
  }}
  var docLinks = [];
  if (inc.source) {{ docLinks.push('<a class="doc-link" href="'+encodeURIComponent(inc.source)+'" target="_blank">'+T('openInvestigationAct')+'</a>'); }}
  if (inc.remediation && inc.remediation.length) {{
    for (var w2=0;w2<inc.remediation.length;w2++) {{
      var wk2 = inc.remediation[w2];
      if (wk2.work_source) {{
        docLinks.push('<a class="doc-link" href="'+encodeURIComponent(wk2.work_source)+'" target="_blank">'+T('openWorkAct')+'</a>');
      }}
    }}
  }}
  if (inc.defect_source) {{ docLinks.push('<a class="doc-link" href="'+encodeURIComponent(inc.defect_source)+'" target="_blank">'+T('openDefectAct')+'</a>'); }}
  if (inc.work_source) {{ docLinks.push('<a class="doc-link" href="'+encodeURIComponent(inc.work_source)+'" target="_blank">'+T('openWorkAct')+'</a>'); }}
  if (docLinks.length) {{ htmlStr += '<p><b>'+T('docs')+'</b> '+docLinks.join(' &middot; ')+'</p>'; }}
  htmlStr += '</div>';
  card.innerHTML = htmlStr;
  (function(cardEl){{
    var btn = cardEl.querySelector('.toggle-btn');
    var det = cardEl.querySelector('.details');
    btn.addEventListener('click', function(){{
      var isOpen = det.className.indexOf('open')!==-1;
      if (isOpen) {{ det.className='details'; btn.innerHTML=T('showDetails'); }}
      else {{ det.className='details open'; btn.innerHTML=T('hideDetails'); }}
    }});
  }})(card);
  return card;
}}

function tokenCoverage(inc, tokens) {{
  if (!tokens.length) return 1;
  var tagToks = [], wordToks = [];
  for (var i=0;i<tokens.length;i++) {{
    if (isTagToken(tokens[i])) {{ tagToks.push(tokens[i]); }} else {{ wordToks.push(tokens[i]); }}
  }}
  var causeText = normalize(inc.cause);
  var allText = normalize([inc.name, inc.cause, inc.category, inc.gpa.join(' '), inc.measures, inc.conclusion, inc.recommendation, inc.prevention_measures||''].join(' '));
  var normTags = (inc.tags||[]).map(function(x){{ return normalize(x); }});
  function coverageFor(toks, isTags) {{
    if (!toks.length) return null;
    var causeMatched = 0, allMatched = 0, tagMatched = 0;
    for (var j=0;j<toks.length;j++) {{
      if (causeText.indexOf(toks[j])!==-1) {{ causeMatched++; }}
      if (allText.indexOf(toks[j])!==-1) {{ allMatched++; }}
      if (isTags) {{
        for (var nt=0;nt<normTags.length;nt++) {{
          if (normTags[nt].indexOf(toks[j])!==-1 || toks[j].indexOf(normTags[nt])!==-1) {{ tagMatched++; break; }}
        }}
      }}
    }}
    if (isTags && normTags.length) {{
      return (tagMatched/toks.length)*0.6 + (causeMatched/toks.length)*0.3 + (allMatched/toks.length)*0.1;
    }}
    return (causeMatched/toks.length)*0.75 + (allMatched/toks.length)*0.25;
  }}
  var tagCov = coverageFor(tagToks, true);
  var wordCov = coverageFor(wordToks, false);
  if (arguments.length>2 && arguments[2]) {{
    if (tagCov!==null && wordCov!==null) {{ return Math.max(tagCov, wordCov); }}
    return tagCov!==null ? tagCov : (wordCov!==null ? wordCov : 0);
  }}
  if (tagCov!==null) {{ return tagCov; }}
  if (wordCov!==null) {{ return wordCov; }}
  return 0;
}}
function buildRecommendation(scored) {{
  if (!scored.length) return '';
  var top = scored[0].inc;
  var top3 = scored.slice(0, Math.min(3, scored.length));
  var catCounts = {{}}, catOrder = [];
  for (var i=0;i<top3.length;i++) {{
    var c = top3[i].inc.category;
    if (catCounts[c]===undefined) {{ catCounts[c]=0; catOrder.push(c); }}
    catCounts[c] = catCounts[c] + 1;
  }}
  var repeatCat = null, repeatCount = 0;
  for (var j=0;j<catOrder.length;j++) {{
    if (catCounts[catOrder[j]] > repeatCount) {{ repeatCount = catCounts[catOrder[j]]; repeatCat = catOrder[j]; }}
  }}
  var pctRounded = Math.round(100*scored[0].pct);
  var measuresText = '';
  if (top.remediation && top.remediation.length) {{
    var parts = [];
    for (var w=0; w<top.remediation.length; w++) {{ parts.push(top.remediation[w].work_description || ''); }}
    measuresText = parts.join('; ');
  }} else {{
    measuresText = top.measures || '';
  }}
  var html = '<div class="rec-block">';
  html += '<div class="rec-title">'+T('recTitle')+'</div>';
  html += '<p class="rec-line"><b>'+T('recMostSimilar')+'</b> '+T('recAct')+' '+escapeHtml(top.act)+' '+T('recFrom')+' '+escapeHtml(top.date)+' &middot; '+T('recMatch')+' '+pctRounded+'% &middot; '+escapeHtml(trCat(top.category))+'</p>';
  if (measuresText) {{
    html += '<p class="rec-line"><b>'+T('recWhatToDo')+'</b> '+escapeHtml(measuresText)+'</p>';
  }}
  if (top.recommendation) {{
    html += '<p class="rec-line"><b>'+T('recFormalRec')+'</b> '+escapeHtml(top.recommendation)+'</p>';
  }}
  if (repeatCat && repeatCount >= 2) {{
    html += '<p class="rec-note">&#9888; ' + FS(
      'Причина &laquo;'+escapeHtml(trCat(repeatCat))+'&raquo; встречается в '+repeatCount+' из '+top3.length+' лучших совпадений — похоже на повторяющуюся системную проблему, стоит проверить не только текущий останов, но и провести системную диагностику.',
      'Cause &laquo;'+escapeHtml(trCat(repeatCat))+'&raquo; appears in '+repeatCount+' of the top '+top3.length+' matches — this looks like a recurring systemic issue. Consider a broader technical check, not just fixing this one shutdown.'
    ) + '</p>';
  }}
  html += '</div>';
  return html;
}}
function render() {{
  var q = document.getElementById('q').value;
  FORCED_TAGS = extractUpperAcronyms(q);
  var tokens = tokenize(q);
  var fGpa = document.getElementById('fGpa').value;
  var fCat = document.getElementById('fCat').value;
  var fYear = document.getElementById('fYear').value;
  function runPass(allowWords) {{
    var out = [];
    for (var i=0;i<incidents.length;i++) {{
      var inc = incidents[i];
      var score = scoreIncident(inc, tokens, allowWords);
      if (score<=0) continue;
      if (fGpa && !arrayContains(inc.gpa, fGpa)) continue;
      if (fCat && inc.category!==fCat) continue;
      if (fYear && inc.date.slice(-4)!==fYear) continue;
      var pct = tokenCoverage(inc, tokens, allowWords);
      if (tokens.length && pct < 0.25) continue;
      out.push({{inc:inc, score:score, pct:pct}});
    }}
    return out;
  }}
  var scored = runPass(false);
  var usedFallback = false;
  if (tokens.length && scored.length===0) {{
    scored = runPass(true);
    usedFallback = true;
  }}
  scored.sort(function(a,b){{ if (b.pct!==a.pct) {{ return b.pct-a.pct; }} return b.score-a.score; }});
  document.getElementById('count').innerHTML = FS('Найдено: ','Found: ')+'<b>'+scored.length+'</b> '+FS('из','of')+' '+incidents.length+(tokens.length?FS(' (схожесть от 100% до 25%)',' (25\u2013100% match)')+(usedFallback?FS(' \u00b7 по смыслу, точных тегов не найдено',' \u00b7 no exact tags, matched by meaning'):''):'');
  var recBox = document.getElementById('recBlock');
  recBox.innerHTML = (tokens.length && scored.length) ? buildRecommendation(scored) : '';
  var box = document.getElementById('results');
  box.innerHTML = '';
  if (scored.length===0) {{
    box.innerHTML = '<div class="empty"><div>&empty;</div>'+T('nothingFound')+'</div>';
    return;
  }}
  for (var s=0;s<scored.length;s++) {{
    var badge = tokens.length ? (FS('Схожесть ','Match ')+Math.round(100*scored[s].pct)+'%') : null;
    box.appendChild(buildCard(scored[s].inc, tokens, badge));
  }}
}}

document.getElementById('q').addEventListener('input', render);
document.getElementById('fGpa').addEventListener('change', render);
document.getElementById('fCat').addEventListener('change', render);
document.getElementById('fYear').addEventListener('change', render);
document.getElementById('reset').addEventListener('click', function(){{
  document.getElementById('q').value='';
  document.getElementById('fGpa').value='';
  document.getElementById('fCat').value='';
  document.getElementById('fYear').value='';
  setFileStatus('');
  render();
}});

function loadScriptOnce(url, cb) {{
  var key = '_loaded_' + url;
  if (window[key]) {{ cb(); return; }}
  var s = document.createElement('script');
  s.src = url;
  s.onload = function(){{ window[key] = true; cb(); }};
  s.onerror = function(){{ cb(new Error('load-failed')); }};
  document.head.appendChild(s);
}}
function setFileStatus(msg, isError) {{
  var el = document.getElementById('fileStatus');
  el.textContent = msg;
  el.className = isError ? 'error' : '';
}}
function extractCircumstancesSection(text) {{
  var m = text.search(/обстоятельств/i);
  if (m === -1) return null;
  var after = text.slice(m);
  var m1 = after.match(/\\)\\s*[:.]/i);
  if (m1 && m1.index < 400) {{
    after = after.slice(m1.index + m1[0].length);
  }} else {{
    var m2 = after.match(/:\\s*\\n/i);
    if (m2 && m2.index < 300) {{ after = after.slice(m2.index + m2[0].length); }}
  }}
  var endMatch = after.search(/\\n\\s*7\\s*[.\\)]|поврежденн|повреждённ|damaged units/i);
  var section = (endMatch > 20) ? after.slice(0, endMatch) : after.slice(0, 1500);
  section = section.trim();
  return section.length > 15 ? section : null;
}}
function extractTagsFromText(text) {{
  var found = [];
  var seen = {{}};
  var tagRe = /\\b[A-ZА-Я]{{1,6}}[-\\s]?\\d{{1,5}}(?:[-\\/][A-Za-z0-9]{{1,6}})*\\b/g;
  var m;
  while ((m = tagRe.exec(text)) !== null) {{
    var raw = m[0].trim();
    var low = raw.toLowerCase().replace(/\\s+/g,'').replace(/-/g,'');
    if (/^(гпа|гпэс|адэс|ибп|вк|ак|мг|кс|гг|зру)№?\\d{{0,2}}$/.test(low)) continue;
    if (low.length < 4) continue;
    if (seen[low]) continue;
    seen[low] = true;
    found.push(raw);
  }}
  var acroRe = /\\b[A-Z]{{3,8}}\\b/g;
  while ((m = acroRe.exec(text)) !== null) {{
    var w = m[0];
    if (ACRONYM_STOPWORDS[w.toLowerCase()]) continue;
    var lw = w.toLowerCase();
    if (seen[lw]) continue;
    seen[lw] = true;
    found.push(w);
  }}
  return found;
}}
function appendReportText(text) {{
  var cleaned = (text || '').replace(/[ \\t]+/g, ' ').replace(/\\n{{3,}}/g, '\\n\\n').trim();
  if (!cleaned) {{ setFileStatus(FS('В файле не найдено текста для поиска.', 'No searchable text found in the file.'), true); return; }}
  var circ = extractCircumstancesSection(cleaned);
  var scopeText = circ && circ.length > 30 ? circ : cleaned.slice(0, 3000);
  var foundTags = extractTagsFromText(scopeText);
  if (!foundTags.length) {{ foundTags = extractTagsFromText(cleaned.slice(0, 4500)); }}
  var finalText;
  if (foundTags.length) {{
    finalText = foundTags.join(' ');
    document.getElementById('q').value = finalText;
    render();
    setFileStatus(FS('Готово — из файла распознаны теги (', 'Done — recognized tags from the file (') + foundTags.join(', ') + FS('), поиск выполнен по ним.', '), searched using them.'));
  }} else {{
    finalText = (circ && circ.length > 30) ? circ + '\\n' + cleaned.slice(0, 2500) : cleaned;
    if (finalText.length > 4500) {{ finalText = finalText.slice(0, 4500); }}
    document.getElementById('q').value = finalText;
    render();
    setFileStatus(FS('Теговых номеров в файле не найдено — показаны похожие случаи по смыслу текста.', 'No tag numbers found in the file — showing similar cases matched by meaning.'));
  }}
}}
function looksGarbled(text) {{
  var norm = normalize(text);
  var raw = norm.split(/\\s+/);
  var commonRu = {{'и':1,'в':1,'на':1,'что':1,'как':1,'при':1,'для':1,'это':1,'по':1,'из':1,'от':1,'до':1,'не':1,'к':1,'с':1,'о':1,'об':1,'же':1,'то':1,'акт':1,'года':1}};
  var found = 0;
  for (var i=0;i<raw.length;i++) {{ if (commonRu[raw[i]]) {{ found++; }} }}
  return raw.length > 15 && found < 2;
}}
function ocrPdfPages(pdf) {{
  var maxPages = Math.min(pdf.numPages, 8);
  setFileStatus(FS('В PDF нет текстового слоя (это скан) — распознаём как изображение, страниц: ', 'The PDF has no text layer (it is a scan) — recognizing as an image, pages: ') + maxPages + FS('. Это может занять пару минут при первой загрузке...', '. This may take a couple of minutes on first load...'));
  loadScriptOnce('https://cdn.jsdelivr.net/npm/tesseract.js@5/dist/tesseract.min.js', function(err){{
    if (err) {{ setFileStatus(FS('Не удалось загрузить библиотеку распознавания текста (нет сети?).', 'Failed to load the text-recognition library (no network?).'), true); return; }}
    window.Tesseract.createWorker(['rus','eng']).then(function(worker){{
      var allText = [];
      function processPage(p) {{
        if (p > maxPages) {{
          worker.terminate();
          var combined = allText.join('\\n');
          if (combined.replace(/\\s+/g,'').length < 15) {{
            setFileStatus(FS('Не удалось распознать текст на странице(ах) — качество скана слишком низкое. Попробуйте вписать обстоятельства вручную в поле поиска.', 'Could not recognize text on the page(s) — scan quality is too low. Try typing the circumstances into the search field manually.'), true);
            return;
          }}
          appendReportText(combined);
          return;
        }}
        setFileStatus(FS('Распознаём страницу ', 'Recognizing page ') + p + FS(' из ', ' of ') + maxPages + '...');
        pdf.getPage(p).then(function(page){{
          var viewport = page.getViewport({{scale: 3}});
          var canvas = document.createElement('canvas');
          canvas.width = viewport.width;
          canvas.height = viewport.height;
          var ctx = canvas.getContext('2d');
          page.render({{canvasContext: ctx, viewport: viewport}}).promise.then(function(){{
            worker.recognize(canvas).then(function(result){{
              allText.push(result.data.text);
              processPage(p + 1);
            }}).catch(function(){{ processPage(p + 1); }});
          }}).catch(function(){{ processPage(p + 1); }});
        }}).catch(function(){{ processPage(p + 1); }});
      }}
      processPage(1);
    }}).catch(function(e){{ setFileStatus(FS('Не удалось запустить распознавание текста: ', 'Failed to start text recognition: ') + e.message, true); }});
  }});
}}
function handleReportFile(file) {{
  try {{
  if (!file) return;
  var name = file.name.toLowerCase();
  setFileStatus(FS('Обработка файла «', 'Processing file "') + file.name + FS('»...', '"...'));
  if (name.endsWith('.txt')) {{
    var r = new FileReader();
    r.onload = function(){{ appendReportText(r.result); }};
    r.onerror = function(){{ setFileStatus(FS('Не удалось прочитать файл.', 'Failed to read the file.'), true); }};
    r.readAsText(file);
  }} else if (name.endsWith('.pdf')) {{
    loadScriptOnce('https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js', function(err){{
      if (err) {{ setFileStatus(FS('Не удалось загрузить библиотеку для чтения PDF (нет сети?).', 'Failed to load the PDF-reading library (no network?).'), true); return; }}
      var pdfjsLib = window.pdfjsLib || window['pdfjs-dist/build/pdf'];
      if (!pdfjsLib) {{ setFileStatus(FS('Библиотека для чтения PDF загрузилась некорректно. Обновите страницу и попробуйте снова.', 'The PDF-reading library loaded incorrectly. Refresh the page and try again.'), true); return; }}
      window.pdfjsLib = pdfjsLib;
      pdfjsLib.GlobalWorkerOptions.workerSrc = 'https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js';
      var reader = new FileReader();
      reader.onload = function(){{
        var typedarray = new Uint8Array(reader.result);
        pdfjsLib.getDocument({{data: typedarray}}).promise.then(function(pdf){{
          var pagePromises = [];
          for (var p = 1; p <= pdf.numPages; p++) {{
            pagePromises.push(pdf.getPage(p).then(function(page){{
              return page.getTextContent().then(function(tc){{
                var lastY = null, parts = [];
                for (var idx=0; idx<tc.items.length; idx++) {{
                  var item = tc.items[idx];
                  var y = (item.transform && item.transform.length>5) ? item.transform[5] : null;
                  if (lastY !== null && y !== null && Math.abs(y - lastY) > 2) {{ parts.push('\\n'); }}
                  parts.push(item.str);
                  parts.push(' ');
                  if (y !== null) {{ lastY = y; }}
                }}
                return parts.join('');
              }});
            }}));
          }}
          Promise.all(pagePromises).then(function(pagesText){{
            var full = pagesText.join('\\n');
            if (full.replace(/\\s+/g,'').length < 15 || looksGarbled(full)) {{
              ocrPdfPages(pdf);
            }} else {{
              appendReportText(full);
            }}
          }}).catch(function(e){{ setFileStatus(FS('Не удалось прочитать текст со страниц PDF: ', 'Failed to read text from the PDF pages: ') + e.message, true); }});
        }}).catch(function(e){{ setFileStatus(FS('Не удалось разобрать PDF-файл: ', 'Failed to parse the PDF file: ') + e.message, true); }});
      }};
      reader.onerror = function(){{ setFileStatus(FS('Не удалось прочитать файл с диска.', 'Failed to read the file from disk.'), true); }};
      reader.readAsArrayBuffer(file);
    }});
  }} else if (name.endsWith('.docx')) {{
    loadScriptOnce('https://cdnjs.cloudflare.com/ajax/libs/mammoth/1.11.0/mammoth.browser.min.js', function(err){{
      if (err) {{ setFileStatus(FS('Не удалось загрузить библиотеку для чтения Word (нет сети?).', 'Failed to load the Word-reading library (no network?).'), true); return; }}
      var reader = new FileReader();
      reader.onload = function(){{
        window.mammoth.extractRawText({{arrayBuffer: reader.result}}).then(function(result){{
          appendReportText(result.value);
        }}).catch(function(e){{ setFileStatus(FS('Не удалось разобрать файл Word: ', 'Failed to parse the Word file: ') + e.message, true); }});
      }};
      reader.onerror = function(){{ setFileStatus(FS('Не удалось прочитать файл с диска.', 'Failed to read the file from disk.'), true); }};
      reader.readAsArrayBuffer(file);
    }});
  }} else if (name.match(/\\.(png|jpe?g)$/)) {{
    setFileStatus(FS('Распознаём текст на фото — это может занять минуту (загружается модель распознавания)...', 'Recognizing text in the photo — this may take a minute (loading the recognition model)...'));
    loadScriptOnce('https://cdn.jsdelivr.net/npm/tesseract.js@5/dist/tesseract.min.js', function(err){{
      if (err) {{ setFileStatus(FS('Не удалось загрузить библиотеку распознавания текста (нет сети?).', 'Failed to load the text-recognition library (no network?).'), true); return; }}
      window.Tesseract.createWorker(['rus','eng']).then(function(worker){{
        worker.recognize(file).then(function(result){{
          var txt = result.data.text || '';
          worker.terminate();
          if (txt.replace(/\\s+/g,'').length < 15) {{
            setFileStatus(FS('Не удалось распознать текст на фото — качество снимка слишком низкое. Попробуйте сфотографировать чётче или вписать обстоятельства вручную.', 'Could not recognize text in the photo — image quality is too low. Try a sharper photo or type the circumstances manually.'), true);
            return;
          }}
          appendReportText(txt);
        }}).catch(function(e){{ setFileStatus(FS('Не удалось распознать текст на фото: ', 'Failed to recognize text in the photo: ') + e.message, true); worker.terminate(); }});
      }}).catch(function(e){{ setFileStatus(FS('Не удалось запустить распознавание текста: ', 'Failed to start text recognition: ') + e.message, true); }});
    }});
  }} else {{
    setFileStatus(FS('Формат не поддерживается. Загрузите PDF, DOCX, TXT, JPG или PNG.', 'Unsupported format. Please upload a PDF, DOCX, TXT, JPG, or PNG.'), true);
  }}
  }} catch (e) {{
    setFileStatus(FS('Непредвиденная ошибка при обработке файла: ', 'Unexpected error while processing the file: ') + e.message, true);
  }}
}}
document.getElementById('reportFile').addEventListener('change', function(e){{
  handleReportFile(e.target.files[0]);
}});

function switchTab(tab) {{
  var isSearch = tab === 'search';
  document.getElementById('searchView').style.display = isSearch ? '' : 'none';
  document.getElementById('dashboardView').style.display = isSearch ? 'none' : '';
  document.getElementById('tabSearchBtn').classList.toggle('active', isSearch);
  document.getElementById('tabDashBtn').classList.toggle('active', !isSearch);
  if (!isSearch) {{ renderDashboard(); }}
}}
document.getElementById('tabSearchBtn').addEventListener('click', function(){{ switchTab('search'); }});
document.getElementById('tabDashBtn').addEventListener('click', function(){{ switchTab('dash'); }});

function applyLanguage() {{
  document.getElementById('pageTitle').textContent = T('title');
  document.getElementById('pageSubtitle').innerHTML = T('subtitle');
  document.getElementById('tabSearchBtn').innerHTML = T('tabSearch');
  document.getElementById('tabDashBtn').innerHTML = T('tabDash');
  document.getElementById('q').placeholder = T('searchPlaceholder');
  document.getElementById('fGpaAll').textContent = T('allGpa');
  document.getElementById('fCatAll').textContent = T('allCat');
  document.getElementById('fYearAll').textContent = T('allYear');
  document.getElementById('reset').textContent = T('reset');
  document.getElementById('uploadLabel').innerHTML = T('uploadLabel');
  document.getElementById('sortLabel').textContent = T('sortRelevance');
  document.getElementById('srcNote').innerHTML = T('srcNote') + (BUILD_STAMP ? ' &middot; ' + (LANG==='ru'?'сборка':'build') + ': ' + BUILD_STAMP : '');
  document.getElementById('langToggle').textContent = LANG==='ru' ? 'EN' : 'RU';
  document.getElementById('langToggle2').textContent = LANG==='ru' ? 'EN' : 'RU';
  document.documentElement.lang = LANG;
  var gpaOpts = document.querySelectorAll('#fGpa option[value]:not([value=""])');
  for (var i=0;i<gpaOpts.length;i++) {{ gpaOpts[i].textContent = trGpa(gpaOpts[i].value); }}
  var catOpts = document.querySelectorAll('#fCat option[value]:not([value=""])');
  for (var j=0;j<catOpts.length;j++) {{ catOpts[j].textContent = trCat(catOpts[j].value); }}
  render();
  _dashRendered = false;
  if (document.getElementById('dashboardView').style.display !== 'none') {{ renderDashboard(); }}
}}
document.getElementById('langToggle').addEventListener('click', function(){{
  LANG = (LANG==='ru') ? 'en' : 'ru';
  applyLanguage();
}});
document.getElementById('langToggle2').addEventListener('click', function(){{
  LANG = (LANG==='ru') ? 'en' : 'ru';
  applyLanguage();
}});

function countBy(list, getterFn) {{
  var map = {{}}, order = [];
  for (var i=0;i<list.length;i++) {{
    var v = getterFn(list[i]);
    if (!(v instanceof Array)) v = [v];
    for (var j=0;j<v.length;j++) {{
      var lbl = v[j];
      if (lbl===undefined || lbl===null || lbl==='') continue;
      if (map[lbl]===undefined) {{ order.push(lbl); map[lbl]=0; }}
      map[lbl] = map[lbl] + 1;
    }}
  }}
  var arr = [];
  for (var k=0;k<order.length;k++) {{ arr.push({{label: order[k], value: map[order[k]]}}); }}
  arr.sort(function(a,b){{ return b.value - a.value; }});
  return arr;
}}
function barRowHtml(item, maxVal, color, translateFn) {{
  var pct = maxVal>0 ? Math.max(2, Math.round(100*item.value/maxVal)) : 2;
  var label = translateFn ? translateFn(item.label) : item.label;
  return '<div class="bar-row"><div class="bar-label" title="'+escapeHtml(label)+'">'+escapeHtml(label)+'</div>' +
    '<div class="bar-track"><div class="bar-fill" style="width:'+pct+'%;background:'+color+';"></div></div>' +
    '<div class="bar-val">'+item.value+'</div></div>';
}}
function chartCardHtml(title, data, color, limit, translateFn) {{
  var top = limit ? data.slice(0, limit) : data;
  var maxVal = top.length ? top[0].value : 0;
  var html = '<div class="chart-card"><h4>'+escapeHtml(title)+'</h4>';
  if (!top.length) {{ html += '<div style="color:var(--text-muted);font-size:13px;">'+T('dashNoData')+'</div>'; }}
  for (var i=0;i<top.length;i++) {{ html += barRowHtml(top[i], maxVal, color, translateFn); }}
  html += '</div>';
  return html;
}}
function yearChartHtml(data) {{
  var sorted = data.slice().sort(function(a,b){{ return a.label < b.label ? -1 : (a.label > b.label ? 1 : 0); }});
  var maxVal = 0;
  for (var i=0;i<sorted.length;i++) {{ if (sorted[i].value>maxVal) maxVal = sorted[i].value; }}
  var html = '<div class="chart-card"><h4>'+T('dashYearChart')+'</h4><div class="yearbars">';
  for (var i=0;i<sorted.length;i++) {{
    var h = maxVal>0 ? Math.max(6, Math.round(100*sorted[i].value/maxVal)) : 6;
    html += '<div class="yearbar-col"><div class="yearbar-val">'+sorted[i].value+'</div>' +
      '<div class="yearbar" style="height:'+h+'%;"></div>' +
      '<div class="yearbar-lbl">'+escapeHtml(sorted[i].label)+'</div></div>';
  }}
  html += '</div></div>';
  return html;
}}
function parseDMY(dateStr) {{
  var parts = (dateStr||'').split('.');
  if (parts.length<3) return new Date(0);
  return new Date(parseInt(parts[2],10), parseInt(parts[1],10)-1, parseInt(parts[0],10));
}}
function computeRepeatRateByYear(accs) {{
  var sorted = accs.slice().sort(function(a,b){{ return parseDMY(a.date) - parseDMY(b.date); }});
  var seenCat = {{}}, yearTotal = {{}}, yearRepeat = {{}}, order = [];
  for (var i=0;i<sorted.length;i++) {{
    var inc = sorted[i];
    var yr = inc.date.slice(-4);
    if (yearTotal[yr]===undefined) {{ yearTotal[yr]=0; yearRepeat[yr]=0; order.push(yr); }}
    yearTotal[yr]++;
    if (seenCat[inc.category]) {{ yearRepeat[yr]++; }}
    seenCat[inc.category] = (seenCat[inc.category]||0) + 1;
  }}
  var arr = [];
  for (var j=0;j<order.length;j++) {{
    var yr2 = order[j];
    var pct = yearTotal[yr2]>0 ? Math.round(100*yearRepeat[yr2]/yearTotal[yr2]) : 0;
    arr.push({{label: yr2, value: pct, total: yearTotal[yr2], repeat: yearRepeat[yr2]}});
  }}
  arr.sort(function(a,b){{ return a.label < b.label ? -1 : (a.label > b.label ? 1 : 0); }});
  return arr;
}}
function percentYearChartHtml(data, title) {{
  var html = '<div class="chart-card"><h4>'+escapeHtml(title)+'</h4><div class="yearbars">';
  if (!data.length) {{ html += '<div style="color:var(--text-muted);font-size:13px;">'+T('dashNoData')+'</div>'; }}
  for (var i=0;i<data.length;i++) {{
    var h = Math.max(4, data[i].value);
    html += '<div class="yearbar-col"><div class="yearbar-val">'+data[i].value+'%</div>' +
      '<div class="yearbar" style="height:'+h+'%;background:var(--coral);" title="'+data[i].repeat+' '+FS('из','of')+' '+data[i].total+'"></div>' +
      '<div class="yearbar-lbl">'+escapeHtml(data[i].label)+'</div></div>';
  }}
  html += '</div></div>';
  return html;
}}
function computeRepeatRegistry(accs) {{
  var byCat = {{}}, order = [];
  for (var i=0;i<accs.length;i++) {{
    var c = accs[i].category;
    if (byCat[c]===undefined) {{ byCat[c] = []; order.push(c); }}
    byCat[c].push(accs[i]);
  }}
  var rows = [];
  for (var j=0;j<order.length;j++) {{
    var list = byCat[order[j]];
    if (list.length < 2) continue;
    list.sort(function(a,b){{ return parseDMY(a.date) - parseDMY(b.date); }});
    rows.push({{category: order[j], count: list.length, first: list[0].date, last: list[list.length-1].date}});
  }}
  rows.sort(function(a,b){{ return b.count - a.count; }});
  return rows;
}}
function repeatRegistryHtml(rows) {{
  var html = '<div class="chart-card"><h4>'+T('dashRegistryTitle')+'</h4>';
  if (!rows.length) {{ html += '<div style="color:var(--text-muted);font-size:13px;">'+T('dashRegistryEmpty')+'</div>'; }}
  for (var i=0;i<rows.length;i++) {{
    var r = rows[i];
    html += '<div class="reg-row"><div class="reg-cat">'+escapeHtml(trCat(r.category))+'</div>' +
      '<div class="reg-count">'+r.count+' '+T('times')+'</div>' +
      '<div class="reg-dates">'+escapeHtml(r.first)+' &rarr; '+escapeHtml(r.last)+'</div></div>';
  }}
  html += '</div>';
  return html;
}}
var _dashRendered = false;
function computeIntervalStats(datesSorted) {{
  var diffs = [];
  for (var i=1;i<datesSorted.length;i++) {{ diffs.push((datesSorted[i]-datesSorted[i-1])/86400000); }}
  if (!diffs.length) return null;
  var sum=0; for (var j=0;j<diffs.length;j++) {{ sum += diffs[j]; }}
  var avg = sum/diffs.length;
  var min = diffs[0], max = diffs[0];
  for (var k=0;k<diffs.length;k++) {{ if (diffs[k]<min) min=diffs[k]; if (diffs[k]>max) max=diffs[k]; }}
  return {{avg:avg, min:min, max:max, n:diffs.length}};
}}
function riskLevel(score) {{
  if (score>=60) return {{text:T('riskHigh'), color:'var(--coral)'}};
  if (score>=35) return {{text:T('riskMedium'), color:'var(--amber)'}};
  return {{text:T('riskLow'), color:'var(--teal)'}};
}}
function computeRiskForecast(accs) {{
  var now = new Date();
  var allDates = [];
  for (var i=0;i<accs.length;i++) {{ allDates.push(parseDMY(accs[i].date)); }}
  allDates.sort(function(a,b){{ return a-b; }});
  var globalStats = computeIntervalStats(allDates);
  var daysSinceGlobalLast = allDates.length ? Math.round((now - allDates[allDates.length-1])/86400000) : null;

  var gpaMap = {{}}, gpaOrder = [];
  for (var j=0;j<accs.length;j++) {{
    var gpas = accs[j].gpa || [];
    for (var g=0; g<gpas.length; g++) {{
      var gp = gpas[g];
      if (gpaMap[gp]===undefined) {{ gpaMap[gp] = []; gpaOrder.push(gp); }}
      gpaMap[gp].push(accs[j]);
    }}
  }}
  var items = [];
  for (var o=0;o<gpaOrder.length;o++) {{
    var gp2 = gpaOrder[o];
    var list = gpaMap[gp2].slice().sort(function(a,b){{ return parseDMY(a.date)-parseDMY(b.date); }});
    var dates = [];
    for (var d=0; d<list.length; d++) {{ dates.push(parseDMY(list[d].date)); }}
    var stats = computeIntervalStats(dates);
    var lastDate = dates[dates.length-1];
    var daysSince = Math.round((now - lastDate)/86400000);
    var avgInterval = stats ? stats.avg : (globalStats ? globalStats.avg : 365);
    var overdueRatio = avgInterval>0 ? daysSince/avgInterval : 0;
    var recentCount=0, priorCount=0;
    for (var d2=0; d2<dates.length; d2++) {{
      var age = (now - dates[d2])/86400000;
      if (age<=730) {{ recentCount++; }} else if (age<=1460) {{ priorCount++; }}
    }}
    var trendFactor = recentCount>priorCount ? 1.15 : (recentCount<priorCount ? 0.85 : 1.0);
    var overdueScore = Math.min(1, overdueRatio) * 50;
    var freqScore = Math.min(1, (list.length/accs.length) * 3) * 30;
    var trendScore = (trendFactor-1) * 100;
    var score = Math.max(5, Math.min(95, Math.round(overdueScore + freqScore + trendScore)));
    var catCounts = countBy(list, function(inc){{ return inc.category; }});
    var topCat = catCounts.length ? catCounts[0] : null;
    items.push({{
      gpa: gp2, count: list.length, daysSince: daysSince, avgInterval: Math.round(avgInterval),
      score: score, topCategory: topCat ? topCat.label : null, topCategoryCount: topCat ? topCat.value : 0
    }});
  }}
  items.sort(function(a,b){{ return b.score - a.score; }});
  return {{items: items, globalStats: globalStats, daysSinceGlobalLast: daysSinceGlobalLast}};
}}
function riskForecastHtml(fc) {{
  var html = '<h3 style="margin:22px 0 12px;font-size:15.5px;color:var(--navy);border-top:1px solid var(--border);padding-top:20px;">'+T('dashRiskTitle')+'</h3>';
  html += '<div class="info-callout"><b>'+T('dashRiskHow')+'</b> '+T('dashRiskHowText')+'</div>';
  html += '<div class="chart-card"><h4>'+T('dashRiskScoring')+'</h4>';
  for (var i=0;i<fc.items.length;i++) {{
    var it = fc.items[i];
    var lvl = riskLevel(it.score);
    html += '<div class="bar-row"><div class="bar-label" style="flex-basis:100px;">'+escapeHtml(trGpa(it.gpa))+'</div>' +
      '<div class="bar-track"><div class="bar-fill" style="width:'+it.score+'%;background:'+lvl.color+';"></div></div>' +
      '<div class="bar-val" style="flex-basis:64px;color:'+lvl.color+';">'+lvl.text+'</div></div>';
  }}
  html += '</div><div class="chart-card"><h4>'+T('dashRiskDetail')+'</h4>';
  for (var j=0;j<fc.items.length;j++) {{
    var it2 = fc.items[j];
    html += '<div class="reg-row"><div class="reg-cat">'+escapeHtml(trGpa(it2.gpa))+' &middot; '+T('riskProbableCause')+' '+(it2.topCategory?escapeHtml(trCat(it2.topCategory)):'&mdash;')+'</div>' +
      '<div class="reg-dates">'+T('riskStopAgo')+' '+it2.daysSince+' '+T('riskDaysAgo')+' &middot; '+T('riskAvgInterval')+' '+it2.avgInterval+' '+T('riskDays')+'</div></div>';
  }}
  html += '</div>';
  if (fc.globalStats) {{
    var gs = fc.globalStats;
    var overdue = fc.daysSinceGlobalLast > gs.avg;
    html += '<div class="chart-card"><h4>'+T('dashForecastTitle')+'</h4>' +
      '<p class="rec-line">'+T('forecastAvgLabel')+' <b>'+Math.round(gs.avg)+' '+T('riskDays')+'</b> ('+FS('мин','min')+' '+Math.round(gs.min)+', '+FS('макс','max')+' '+Math.round(gs.max)+', '+FS('на основе '+gs.n+' интервалов','based on '+gs.n+' intervals')+').</p>' +
      '<p class="rec-line">'+T('forecastSinceLabel')+' <b>'+fc.daysSinceGlobalLast+' '+T('riskDays')+'</b> — '+FS('это','which is')+' '+(overdue?T('forecastMore'):T('forecastLess'))+' '+FS('среднего интервала, вероятность нового останова','than the average interval — probability of a new shutdown is')+' '+(overdue?'<b>'+T('forecastElevated')+'</b>':T('forecastNormal'))+'.</p>' +
      '</div>';
  }}
  return html;
}}
function renderDashboard() {{
  if (_dashRendered) return;
  var accs = [];
  for (var i=0;i<incidents.length;i++) {{ if (incidents[i].kind==='incident') accs.push(incidents[i]); }}
  if (!accs.length) {{
    document.getElementById('dashContent').innerHTML = '<div class="empty"><div>&#9888;</div>'+T('dashEmpty')+'</div>';
    return;
  }}
  var byGpa = countBy(accs, function(inc){{ return inc.gpa; }});
  var byYear = countBy(accs, function(inc){{ return inc.date.slice(-4); }});
  var byCategory = countBy(accs, function(inc){{ return inc.category; }});
  var byType = countBy(accs, function(inc){{ return inc.type; }});
  var topGpa = byGpa.length ? byGpa[0] : null;
  var topCat = byCategory.length ? byCategory[0] : null;
  var years = [];
  for (var y=0;y<byYear.length;y++) {{ years.push(byYear[y].label); }}
  var minY = years.length ? years.reduce(function(a,b){{ return a<b?a:b; }}) : '—';
  var maxY = years.length ? years.reduce(function(a,b){{ return a>b?a:b; }}) : '—';

  var html = '<div class="dash-grid">' +
    '<div class="stat-card"><div class="num">'+accs.length+'</div><div class="lbl">'+T('dashStatTotal')+'</div></div>' +
    '<div class="stat-card"><div class="num">'+minY+'&ndash;'+maxY+'</div><div class="lbl">'+T('dashStatPeriod')+'</div></div>' +
    '<div class="stat-card"><div class="num">'+(topGpa?escapeHtml(trGpa(topGpa.label)):'—')+'</div><div class="lbl">'+T('dashStatTopGpa')+' &middot; '+(topGpa?topGpa.value:0)+' '+T('times')+'</div></div>' +
    '<div class="stat-card"><div class="num" style="font-size:14.5px;">'+(topCat?escapeHtml(trCat(topCat.label)):'—')+'</div><div class="lbl">'+T('dashStatTopCat')+' &middot; '+(topCat?topCat.value:0)+' '+T('times')+'</div></div>' +
    '</div>';
  html += yearChartHtml(byYear);
  html += chartCardHtml(T('dashByGpa'), byGpa, 'var(--teal)', null, trGpa);
  html += chartCardHtml(T('dashTopCat'), byCategory, 'var(--navy)', 8, trCat);
  html += chartCardHtml(T('dashByType'), byType, 'var(--coral)', null, trType);

  var repeatByYear = computeRepeatRateByYear(accs);
  var registry = computeRepeatRegistry(accs);
  var totalRepeatIncidents = 0;
  for (var rr=0; rr<repeatByYear.length; rr++) {{ totalRepeatIncidents += repeatByYear[rr].repeat; }}
  var repeatSharePct = accs.length ? Math.round(100*totalRepeatIncidents/accs.length) : 0;

  html += '<h3 style="margin:22px 0 12px;font-size:15.5px;color:var(--navy);border-top:1px solid var(--border);padding-top:20px;">'+T('dashMetricsTitle')+'</h3>';
  html += '<div class="info-callout"><b>'+T('dashGoal')+'</b> '+T('dashGoalText')+'</div>';
  html += '<div class="dash-grid">' +
    '<div class="stat-card"><div class="num">'+repeatSharePct+'%</div><div class="lbl">'+T('dashRepeatShare')+'</div></div>' +
    '<div class="stat-card"><div class="num">'+registry.length+'</div><div class="lbl">'+T('dashRepeatCats')+'</div></div>' +
    '</div>';
  html += percentYearChartHtml(repeatByYear, T('dashRepeatTrend'));
  html += repeatRegistryHtml(registry);

  var riskForecast = computeRiskForecast(accs);
  html += riskForecastHtml(riskForecast);

  document.getElementById('dashContent').innerHTML = html;
  _dashRendered = true;
}}

Promise.all([
  fetch('accidents.json').then(function(r){{ return r.json(); }}),
  fetch('defects.json').then(function(r){{ return r.json(); }}).catch(function(){{ return []; }})
]).then(function(results){{
  var accidents = results[0], defects = results[1];
  incidents = [];
  for (var i=0;i<accidents.length;i++) {{
    var a = accidents[i];
    incidents.push({{
      kind:'incident', act:a.act, date:a.date, time:a.time, gpa:a.gpa, type:a.type,
      category:a.category, name:a.name, cause:a.cause, measures:a.measures,
      conclusion:a.conclusion, recommendation:a.recommendation,
      prevention_measures:a.prevention_measures||'',
      remediation:a.remediation||[], source:a.source, tags:a.tags||[]
    }});
  }}
  for (var j=0;j<defects.length;j++) {{
    var d = defects[j];
    var materialsTxt = '';
    if (d.materials && d.materials.length) {{
      var parts = [];
      for (var m=0;m<d.materials.length;m++) {{
        var mt = d.materials[m];
        parts.push(mt.name+' ('+mt.part_no+') — '+mt.qty+' '+mt.unit);
      }}
      materialsTxt = ' Расходные материалы: '+parts.join('; ');
    }}
    incidents.push({{
      kind:'defect', act:d.defect_act, date:d.defect_date, time:'', gpa:d.gpa||[],
      type:'Дефект (без останова)', category:'Плановая инспекция / дефект',
      name:d.title || (d.description||'').slice(0,80), cause:d.description,
      measures:'', conclusion:'Статус: '+(d.status||''),
      recommendation:(d.work_description||'')+materialsTxt, remediation:[],
      defect_source:d.defect_source, work_source:d.work_source
    }});
  }}
  initFilters();
  applyLanguage();
}}).catch(function(err){{
  document.getElementById('results').innerHTML =
    '<div class="empty"><div>&#9888;</div>' + FS('Не удалось загрузить accidents.json / defects.json.<br>', 'Failed to load accidents.json / defects.json.<br>') +
    FS('Если вы открыли файл напрямую (file://) — это ожидаемо: браузеры блокируют fetch() для локальных файлов.<br>', 'If you opened the file directly (file://) this is expected — browsers block fetch() for local files.<br>') +
    FS('Разместите сайт на GitHub Pages (см. README) или запустите локальный сервер: <code>python3 -m http.server</code>', 'Deploy the site on GitHub Pages (see README) or run a local server: <code>python3 -m http.server</code>') +
    '</div>';
  console.error(err);
}});
</script>
</body>
</html>
"""


def build_web_site(incidents, defects, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    with open(out_dir / 'accidents.json', 'w', encoding='utf-8') as f:
        json.dump(incidents, f, ensure_ascii=False, indent=2)
    with open(out_dir / 'defects.json', 'w', encoding='utf-8') as f:
        json.dump(defects, f, ensure_ascii=False, indent=2)
    # WEB_HTML_TEMPLATE не проходит через .format() (нет плейсхолдеров для подстановки),
    # поэтому двойные скобки {{ }}, унаследованные при написании шаблона, нужно свести
    # к обычным одиночным { } — иначе браузер не распознает CSS/JS.
    fixed_html = WEB_HTML_TEMPLATE.replace('{{', '{').replace('}}', '}')
    build_stamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    fixed_html = fixed_html.replace(
        "var BUILD_STAMP = '';",
        "var BUILD_STAMP = '" + build_stamp + "';"
    )
    with open(out_dir / 'index.html', 'w', encoding='utf-8') as f:
        f.write(fixed_html)


def load_data():
    with open(DATA_FILE, encoding='utf-8') as f:
        incidents = json.load(f)
    for inc in incidents:
        inc['year'] = int(inc['date'].split('.')[-1])
        inc.setdefault('remediation', [])
    return incidents


def load_defects():
    """Дефекты, найденные вне рамок автоостанова (плановые инспекции и т.п.)."""
    if not DEFECTS_FILE.exists():
        return []
    with open(DEFECTS_FILE, encoding='utf-8') as f:
        return json.load(f)


def remediation_text(inc):
    """«Принятые меры»: если есть акт(ы) выполненных работ по устранению — показываем их;
    если нет — используем пункт 8 акта расследования (принятые меры для устранения причин
    останова, поле measures)."""
    works = inc.get('remediation', [])
    if not works:
        return inc.get('measures', '—')
    lines = []
    for w in works:
        status = w.get('status', 'выполнено')
        line = f"{w.get('work_description', '')} — Акт выполненных работ {w.get('work_act', '?')} " \
               f"от {w.get('work_date', '?')} ({status})"
        lines.append(line)
    return '\n'.join(lines)


def remediation_rows(incidents, defects):
    """Единый реестр работ по устранению: и привязанные к авариям, и найденные отдельно (дефекты)."""
    rows = []
    for inc in incidents:
        for w in inc.get('remediation', []):
            rows.append({
                'incident_act': inc['act'], 'incident_date': inc['date'], 'gpa': ', '.join(inc['gpa']),
                'defect_act': w.get('defect_act', ''), 'description': w.get('description', ''),
                'work_description': w.get('work_description', ''),
                'work_act': w.get('work_act', ''), 'work_date': w.get('work_date', ''),
                'status': w.get('status', 'выполнено'),
                'defect_source': w.get('defect_source', w.get('source', '')),
                'work_source': w.get('work_source', ''),
                'materials': w.get('materials', []),
            })
    for d in defects:
        rows.append({
            'incident_act': '— (плановая инспекция)', 'incident_date': d.get('defect_date', ''),
            'gpa': ', '.join(d.get('gpa', [])),
            'defect_act': d.get('defect_act', ''), 'description': d.get('description', ''),
            'work_description': d.get('work_description', ''),
            'work_act': d.get('work_act', ''), 'work_date': d.get('work_date', ''),
            'status': d.get('status', 'выполнено'),
            'defect_source': d.get('defect_source', ''),
            'work_source': d.get('work_source', ''),
            'materials': d.get('materials', []),
        })
    return rows


# =====================================================================
# ЧАСТЬ 1 — генерация xlsx
# =====================================================================
def header_cell(cell, text, size=12, fill='FF1A365D', color='FFFFFFFF'):
    cell.value = text
    cell.font = Font(name=FONT, size=size, bold=True, color=color)
    cell.fill = PatternFill(fill_type='solid', fgColor=fill)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER


def build_xlsx(incidents, defects):
    wb = openpyxl.Workbook()

    # ---- Лист 1: База_данных ----
    ws1 = wb.active
    ws1.title = 'База_данных'
    ws1.merge_cells('A1:F1')
    header_cell(ws1['A1'], 'БАЗА ДАННЫХ ПО АВТО ОСТАНОВАМ ГПА', size=14)
    ws1.row_dimensions[1].height = 28
    headers = ['Год', 'ГПА', 'Кол-во остановов', 'Дата', '№ Акта', 'Источник (файл)']
    for i, h in enumerate(headers, start=1):
        header_cell(ws1.cell(row=3, column=i), h, size=11, fill=GREY, color='FF2D3748')

    r = 4
    for inc in incidents:
        for gpa in inc['gpa']:
            ws1.cell(row=r, column=1, value=inc['year']).font = Font(name=FONT, size=11)
            ws1.cell(row=r, column=2, value=gpa).font = Font(name=FONT, size=11)
            ws1.cell(row=r, column=3, value=1).font = Font(name=FONT, size=11)
            ws1.cell(row=r, column=4, value=inc['date']).font = Font(name=FONT, size=11)
            ws1.cell(row=r, column=5, value=f"{inc['date'][-4:]}-{inc['act']}").font = Font(name=FONT, size=11)
            ws1.cell(row=r, column=6, value=inc['source']).font = Font(name=FONT, size=8, italic=True, color='FF718096')
            for c in range(1, 7):
                ws1.cell(row=r, column=c).border = BORDER
                ws1.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            r += 1
    last_data_row = r - 1

    note_row = r + 1
    ws1.merge_cells(f'A{note_row}:F{note_row}')
    note = ws1.cell(row=note_row, column=1)
    note.value = ('Этот лист формируется автоматически из accidents.json при запуске build.py. '
                  'Не редактируйте его вручную — правки потеряются при следующей пересборке.')
    note.font = Font(name=FONT, size=9, italic=True, color='FFC0392B')
    note.alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[note_row].height = 28

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 13
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 55
    ws1.sheet_view.showGridLines = False

    # ---- Лист 2: Таблица_аварий ----
    ws2 = wb.create_sheet('Таблица_аварий')
    ws2.merge_cells('A1:G1')
    header_cell(ws2['A1'], 'ТАБЛИЦА АВАРИЙНЫХ ОСТАНОВОВ ГПА', size=14)
    ws2.row_dimensions[1].height = 28
    headers2 = ['№', 'Наименование (тема аварии)', 'Дата', '№ акта', 'ГПА', 'Источник', 'Стр. в PDF']
    for i, h in enumerate(headers2, start=1):
        header_cell(ws2.cell(row=3, column=i), h, size=11, fill=GREY, color='FF2D3748')

    # рассчитываем стартовые страницы в объединённом PDF (примерно: нужно
    # пересчитывать merge_acts.py при добавлении нового акта — номера ниже
    # актуальны только для уже смерженных 14 актов)
    page_map = {}
    running_page = 1
    page_counts = [3, 4, 5, 3, 3, 9, 5, 3, 6, 3, 5, 3, 6, 5]
    for idx, inc in enumerate(incidents):
        page_map[inc['source']] = running_page
        running_page += page_counts[idx] if idx < len(page_counts) else 1

    r = 4
    for i, inc in enumerate(incidents, start=1):
        ws2.cell(row=r, column=1, value=i).font = Font(name=FONT, size=11)
        ws2.cell(row=r, column=2, value=inc['name']).font = Font(name=FONT, size=11)
        ws2.cell(row=r, column=3, value=inc['date']).font = Font(name=FONT, size=11)
        ws2.cell(row=r, column=4, value=inc['act']).font = Font(name=FONT, size=11)
        ws2.cell(row=r, column=5, value=', '.join(inc['gpa'])).font = Font(name=FONT, size=11)
        link_cell = ws2.cell(row=r, column=6, value='Открыть акт (PDF)')
        link_cell.hyperlink = 'Все_акты_расследований.pdf'
        link_cell.font = Font(name=FONT, size=11, color='FF1155CC', underline='single')
        ws2.cell(row=r, column=7, value=page_map.get(inc['source'], '')).font = Font(name=FONT, size=11, bold=True)
        for c in range(1, 8):
            ws2.cell(row=r, column=c).border = BORDER
            ws2.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
        ws2.row_dimensions[r].height = 32
        r += 1

    widths2 = [6, 46, 12, 10, 16, 16, 11]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A4'

    # ---- Лист 3: Сводная_таблица (формулы) ----
    ws3 = wb.create_sheet('Сводная_таблица')
    ws3.merge_cells('A1:D1')
    header_cell(ws3['A1'], 'СВОДНАЯ ТАБЛИЦА ПО ГОДАМ (авторасчёт по данным листа "База_данных")', size=13)
    ws3.row_dimensions[1].height = 30
    for i, h in enumerate(['Год', 'ГПА №1', 'ГПА №2', 'ГПА №3'], start=1):
        header_cell(ws3.cell(row=3, column=i), h, size=11, fill=GREY, color='FF2D3748')

    years = sorted(set(inc['year'] for inc in incidents))
    years = list(range(years[0], years[-1] + 1))
    rng_year = f"База_данных!$A$4:$A${last_data_row}"
    rng_gpa = f"База_данных!$B$4:$B${last_data_row}"
    rng_count = f"База_данных!$C$4:$C${last_data_row}"

    for idx, year in enumerate(years):
        row = 4 + idx
        ws3.cell(row=row, column=1, value=year).font = Font(name=FONT, size=11, bold=True)
        for col_idx, gpa_name in enumerate(['ГПА №1', 'ГПА №2', 'ГПА №3'], start=2):
            formula = (f'=IFERROR(IF(SUMIFS({rng_count},{rng_year},A{row},{rng_gpa},"{gpa_name}")=0,"",'
                       f'SUMIFS({rng_count},{rng_year},A{row},{rng_gpa},"{gpa_name}")),"")')
            ws3.cell(row=row, column=col_idx, value=formula).font = Font(name=FONT, size=11)
        for c in range(1, 5):
            ws3.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
            ws3.cell(row=row, column=c).border = BORDER

    total_row = 4 + len(years)
    ws3.cell(row=total_row, column=1, value='ИТОГО').font = Font(name=FONT, size=11, bold=True)
    for col_idx, gpa_name in enumerate(['ГПА №1', 'ГПА №2', 'ГПА №3'], start=2):
        formula = f'=SUMIFS({rng_count},{rng_gpa},"{gpa_name}")'
        cell = ws3.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor=GREY)
    ws3.cell(row=total_row, column=1).fill = PatternFill(fill_type='solid', fgColor=GREY)
    for c in range(1, 5):
        ws3.cell(row=total_row, column=c).alignment = Alignment(horizontal='center', vertical='center')
        ws3.cell(row=total_row, column=c).border = BORDER
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 16
    ws3.sheet_view.showGridLines = False

    # ---- Лист 4: Инциденты_детально ----
    ws4 = wb.create_sheet('Инциденты_детально')
    ws4.merge_cells('A1:I1')
    header_cell(ws4['A1'], 'ДЕТАЛИЗАЦИЯ ИНЦИДЕНТОВ ПО АКТАМ ТЕХНИЧЕСКОГО РАССЛЕДОВАНИЯ', size=13)
    ws4.row_dimensions[1].height = 26
    headers4 = ['№ Акта', 'Дата', 'Время', 'ГПА', 'Тип останова', 'Категория причины',
                'Обстоятельства, при которых произошел останов (п.6)', 'Меры при останове', 'Заключение / Принятые меры']
    for i, h in enumerate(headers4, start=1):
        header_cell(ws4.cell(row=3, column=i), h, size=10, fill=GREY, color='FF2D3748')

    r = 4
    for inc in incidents:
        ws4.cell(row=r, column=1, value=f"{inc['act']} от {inc['date']}").font = Font(name=FONT, size=10, bold=True)
        ws4.cell(row=r, column=2, value=inc['date']).font = Font(name=FONT, size=10)
        ws4.cell(row=r, column=3, value=inc['time']).font = Font(name=FONT, size=10)
        ws4.cell(row=r, column=4, value=', '.join(inc['gpa'])).font = Font(name=FONT, size=10)
        ws4.cell(row=r, column=5, value=inc['type']).font = Font(name=FONT, size=10)
        ws4.cell(row=r, column=6, value=inc['category']).font = Font(name=FONT, size=10)
        ws4.cell(row=r, column=7, value=inc['cause']).font = Font(name=FONT, size=9)
        extra = f"\nПовреждено: {inc['damaged']}" if inc['damaged'] != 'Отсутствуют' else ''
        ws4.cell(row=r, column=8, value=inc['measures'] + extra).font = Font(name=FONT, size=9)
        ws4.cell(row=r, column=9, value=inc['conclusion'] + '\n' + remediation_text(inc)).font = Font(name=FONT, size=9)
        for c in range(1, 10):
            ws4.cell(row=r, column=c).border = BORDER
            ws4.cell(row=r, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        ws4.row_dimensions[r].height = 95 + 18 * len(inc.get('remediation', []))
        r += 1
    widths4 = [16, 11, 8, 14, 14, 20, 45, 35, 35]
    for i, w in enumerate(widths4, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A4'

    # ---- Лист 5: Классификатор_причин (формулы) ----
    ws5 = wb.create_sheet('Классификатор_причин')
    ws5.merge_cells('A1:C1')
    header_cell(ws5['A1'], 'КЛАССИФИКАЦИЯ ПРИЧИН ОТКАЗОВ (авторасчёт)', size=13)
    ws5.row_dimensions[1].height = 26
    for i, h in enumerate(['Категория причины', 'Кол-во инцидентов', 'Доля, %'], start=1):
        header_cell(ws5.cell(row=3, column=i), h, size=11, fill=GREY, color='FF2D3748')

    categories = sorted(set(inc['category'] for inc in incidents))
    r = 4
    first_row = r
    for cat in categories:
        ws5.cell(row=r, column=1, value=cat).font = Font(name=FONT, size=11)
        formula_count = f"=COUNTIF('Инциденты_детально'!$F$4:$F${3+len(incidents)},A{r})"
        ws5.cell(row=r, column=2, value=formula_count).font = Font(name=FONT, size=11)
        ws5.cell(row=r, column=3, value=f'=B{r}/{len(incidents)}').font = Font(name=FONT, size=11, italic=True)
        ws5.cell(row=r, column=3).number_format = '0.0%'
        for c in range(1, 4):
            ws5.cell(row=r, column=c).border = BORDER
            ws5.cell(row=r, column=c).alignment = Alignment(vertical='center', wrap_text=True)
        r += 1
    total_row5 = r
    ws5.cell(row=total_row5, column=1, value='ИТОГО инцидентов').font = Font(name=FONT, size=11, bold=True)
    ws5.cell(row=total_row5, column=2, value=f'=SUM(B{first_row}:B{r-1})').font = Font(name=FONT, size=11, bold=True)
    ws5.cell(row=total_row5, column=3, value=f'=SUM(C{first_row}:C{r-1})').font = Font(name=FONT, size=11, bold=True)
    ws5.cell(row=total_row5, column=3).number_format = '0.0%'
    for c in range(1, 4):
        ws5.cell(row=total_row5, column=c).fill = PatternFill(fill_type='solid', fgColor=GREY)
        ws5.cell(row=total_row5, column=c).border = BORDER
    ws5.column_dimensions['A'].width = 42
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 12
    ws5.sheet_view.showGridLines = False

    # ---- Лист 6: Устранение_дефектов ----
    ws6r = wb.create_sheet('Устранение_дефектов')
    ws6r.merge_cells('A1:J1')
    header_cell(ws6r['A1'], 'РЕЕСТР РАБОТ ПО УСТРАНЕНИЮ ДЕФЕКТОВ', size=13)
    ws6r.row_dimensions[1].height = 26
    headers6r = ['Привязка к аварии (акт/дата)', 'ГПА', '№ дефектного акта', 'Описание дефекта',
                 'Описание выполненных работ', '№ акта вып. работ', 'Дата выполнения', 'Статус',
                 'Дефектный акт', 'Акт вып. работ']
    for i, h in enumerate(headers6r, start=1):
        header_cell(ws6r.cell(row=3, column=i), h, size=10, fill=GREY, color='FF2D3748')

    rows6r = remediation_rows(incidents, defects)
    r = 4
    for row in rows6r:
        ws6r.cell(row=r, column=1, value=f"{row['incident_act']} от {row['incident_date']}").font = Font(name=FONT, size=10)
        ws6r.cell(row=r, column=2, value=row['gpa']).font = Font(name=FONT, size=10)
        ws6r.cell(row=r, column=3, value=row['defect_act']).font = Font(name=FONT, size=10, bold=True)
        ws6r.cell(row=r, column=4, value=row['description']).font = Font(name=FONT, size=9)
        work_desc = row['work_description']
        if row.get('materials'):
            mat_lines = [f"{m['name']} ({m['part_no']}) — {m['qty']} {m['unit']}" for m in row['materials']]
            work_desc += '\nРасходные материалы: ' + '; '.join(mat_lines)
        ws6r.cell(row=r, column=5, value=work_desc).font = Font(name=FONT, size=9)
        ws6r.cell(row=r, column=6, value=row['work_act']).font = Font(name=FONT, size=10, bold=True)
        ws6r.cell(row=r, column=7, value=row['work_date']).font = Font(name=FONT, size=10)
        status_cell = ws6r.cell(row=r, column=8, value=row['status'])
        status_cell.font = Font(name=FONT, size=10, bold=True,
                                 color='FF0F6E56' if row['status'] == 'выполнено' else 'FF993C1D')
        if row['defect_source']:
            link_cell = ws6r.cell(row=r, column=9, value='Открыть дефектный акт')
            link_cell.hyperlink = row['defect_source']
            link_cell.font = Font(name=FONT, size=10, color='FF1155CC', underline='single')
        if row['work_source']:
            link_cell2 = ws6r.cell(row=r, column=10, value='Открыть акт вып. работ')
            link_cell2.hyperlink = row['work_source']
            link_cell2.font = Font(name=FONT, size=10, color='FF1155CC', underline='single')
        for c in range(1, 11):
            ws6r.cell(row=r, column=c).border = BORDER
            ws6r.cell(row=r, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        ws6r.row_dimensions[r].height = 55
        r += 1

    if not rows6r:
        ws6r.merge_cells('A4:J4')
        empty_cell = ws6r.cell(row=4, column=1)
        empty_cell.value = ('Пока нет данных о выполненных работах. Заполняется полями "remediation" в '
                             'accidents.json (для аварий) или записями в defects.json (для дефектов без остановов).')
        empty_cell.font = Font(name=FONT, size=10, italic=True, color='FF718096')
        empty_cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws6r.row_dimensions[4].height = 30

    widths6r = [18, 14, 14, 38, 38, 20, 13, 12, 16, 16]
    for i, w in enumerate(widths6r, start=1):
        ws6r.column_dimensions[get_column_letter(i)].width = w
    ws6r.sheet_view.showGridLines = False
    ws6r.freeze_panes = 'A4'

    # ---- Лист 7: Источники ----
    ws6 = wb.create_sheet('Источники')
    ws6.merge_cells('A1:D1')
    header_cell(ws6['A1'], 'ЖУРНАЛ ИСТОЧНИКОВ ДАННЫХ', size=13)
    ws6.row_dimensions[1].height = 26
    for i, h in enumerate(['№', 'Акт', 'Файл', 'Дата акта'], start=1):
        header_cell(ws6.cell(row=3, column=i), h, size=11, fill=GREY, color='FF2D3748')
    r = 4
    for i, inc in enumerate(incidents, start=1):
        ws6.cell(row=r, column=1, value=i).font = Font(name=FONT, size=10)
        ws6.cell(row=r, column=2, value=inc['act']).font = Font(name=FONT, size=10)
        ws6.cell(row=r, column=3, value=inc['source']).font = Font(name=FONT, size=10)
        ws6.cell(row=r, column=4, value=inc['date']).font = Font(name=FONT, size=10)
        for c in range(1, 5):
            ws6.cell(row=r, column=c).border = BORDER
            ws6.cell(row=r, column=c).alignment = Alignment(vertical='top', wrap_text=True)
        r += 1
    ws6.column_dimensions['A'].width = 6
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 55
    ws6.column_dimensions['D'].width = 14
    ws6.sheet_view.showGridLines = False

    wb.save(XLSX_OUT)
    return last_data_row


# =====================================================================
# ЧАСТЬ 2 — генерация html (данные собираются из того же accidents.json)
# =====================================================================
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta http-equiv="X-UA-Compatible" content="IE=edge">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Поиск по базе авто остановов ГПА</title>
<style>
  :root{{
    --navy:#1A365D;
    --navy-dark:#0F2440;
    --surface:#F7FAFC;
    --card:#FFFFFF;
    --border:#DDE3EA;
    --text:#1F2933;
    --text-muted:#5C6B7A;
    --teal:#0F6E56;
    --teal-bg:#E1F5EE;
    --coral:#993C1D;
    --coral-bg:#FAECE7;
    --amber:#854F0B;
    --amber-bg:#FAEEDA;
    --mono: "SF Mono", "Consolas", "Roboto Mono", monospace;
    --sans: "Segoe UI", Inter, system-ui, -apple-system, Arial, sans-serif;
  }}
  *{{box-sizing:border-box;}}
  body{{
    margin:0; background:var(--surface); color:var(--text);
    font-family:var(--sans); line-height:1.5;
  }}
  header{{
    background:linear-gradient(180deg,var(--navy),var(--navy-dark));
    color:#fff; padding:28px 24px 22px;
  }}
  header h1{{margin:0 0 4px; font-size:20px; font-weight:600;}}
  header p{{margin:0; font-size:13px; color:#C9D6E5;}}

  .container{{max-width:960px; margin:0 auto; padding:20px 20px 60px;}}

  .toolbar{{
    background:var(--card); border:1px solid var(--border); border-radius:10px;
    padding:16px; margin:-18px 0 20px; box-shadow:0 2px 10px rgba(15,36,64,0.08);
  }}
  .search-row{{display:flex; gap:10px; margin-bottom:12px;}}
  .search-row input{{
    flex:1; padding:11px 14px; font-size:14px; border:1px solid var(--border);
    border-radius:8px; outline:none; font-family:var(--sans);
  }}
  .search-row input:focus{{border-color:var(--navy); box-shadow:0 0 0 3px rgba(26,54,93,0.12);}}

  .filters{{display:flex; gap:10px; flex-wrap:wrap;}}
  .filters select{{
    padding:8px 10px; font-size:13px; border:1px solid var(--border);
    border-radius:7px; background:#fff; color:var(--text); font-family:var(--sans);
  }}
  .filters button{{
    padding:8px 14px; font-size:13px; border:1px solid var(--border);
    border-radius:7px; background:var(--surface); color:var(--text-muted);
    cursor:pointer;
  }}
  .filters button:hover{{background:#EDF1F5;}}

  .meta-row{{
    display:flex; justify-content:space-between; align-items:baseline;
    margin-bottom:12px; font-size:13px; color:var(--text-muted);
  }}
  .meta-row b{{color:var(--text);}}

  .card{{
    background:var(--card); border:1px solid var(--border); border-radius:10px;
    padding:16px 18px; margin-bottom:12px;
  }}
  .card-top{{display:flex; gap:8px; align-items:center; margin-bottom:8px; flex-wrap:wrap;}}
  .tag{{
    font-family:var(--mono); font-size:11.5px; padding:3px 8px; border-radius:5px;
    font-weight:600; letter-spacing:0.2px;
  }}
  .tag.act{{background:#EDF1F5; color:var(--navy);}}
  .tag.gpa{{background:var(--teal-bg); color:var(--teal);}}
  .tag.cat{{background:var(--amber-bg); color:var(--amber);}}
  .tag.type{{background:var(--coral-bg); color:var(--coral);}}

  .card h3{{margin:0 0 6px; font-size:15.5px; font-weight:600; color:var(--text);}}
  .card .cause{{font-size:13.5px; color:var(--text-muted); margin:0 0 8px;}}
  mark{{background:#FFE9A8; color:#412402; padding:0 2px; border-radius:2px;}}

  .details{{display:none; font-size:13px; margin-top:8px; padding-top:10px; border-top:1px dashed var(--border);}}
  .details.open{{display:block;}}
  .details p{{margin:0 0 8px;}}
  .details b{{color:var(--text);}}

  .toggle-btn{{
    font-size:12.5px; color:var(--navy); background:none; border:none; cursor:pointer;
    padding:0; font-weight:600; text-decoration:underline;
  }}

  .empty{{text-align:center; color:var(--text-muted); padding:50px 10px; font-size:14px;}}
  .empty div{{font-size:28px; margin-bottom:8px;}}

  .src-note{{ font-size: 11px; color: #9AA5B1; text-align: center; margin-top: 24px; }}

  a.doc-link{{color:var(--navy); font-weight:600; text-decoration:underline;}}

  @media (max-width:600px){{
    .search-row{{flex-direction:column;}}
  }}
</style>
</head>
<body>
<header>
  <h1>Поиск по базе авто остановов ГПА</h1>
  <p>Умная система анализа авто остановов и помощи инженерам &middot; КС-1 &laquo;Алимтау&raquo;</p>
</header>
<div class="container">
  <div class="toolbar">
    <div class="search-row">
      <input id="q" type="text" placeholder="Например: свечной кран, потеря пламени, RB6-2, ГПА№2..." autocomplete="off">
    </div>
    <div class="filters">
      <select id="fGpa"><option value="">Все ГПА</option></select>
      <select id="fCat"><option value="">Все категории</option></select>
      <select id="fYear"><option value="">Все годы</option></select>
      <button id="reset" type="button">Сбросить</button>
    </div>
  </div>
  <div class="meta-row">
    <span class="left" id="count"></span>
    <span class="right">Сортировка по релевантности</span>
  </div>
  <div id="results"></div>
  <div class="src-note">Сгенерировано автоматически из accidents.json + defects.json &middot; версия данных: {n_records} записей</div>
</div>
<script>
var incidents = {incidents_json};

function uniqueSorted(arr) {{
  var seen = {{}}, out = [];
  for (var i = 0; i < arr.length; i++) {{ var v = arr[i]; if (!seen[v]) {{ seen[v] = true; out.push(v); }} }}
  out.sort();
  return out;
}}
function arrayContains(arr, val) {{
  for (var i = 0; i < arr.length; i++) {{ if (arr[i] === val) return true; }}
  return false;
}}
function escapeHtml(s) {{
  return String(s).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
}}
function escapeRegExp(s) {{ return s.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&'); }}

var allGpa = [], allCat = [], allYear = [];
for (var i = 0; i < incidents.length; i++) {{
  var inc = incidents[i];
  for (var j = 0; j < inc.gpa.length; j++) {{ allGpa.push(inc.gpa[j]); }}
  allCat.push(inc.category);
  allYear.push(inc.date.slice(-4));
}}
function fillSelect(id, values) {{
  var sel = document.getElementById(id);
  var uniq = uniqueSorted(values);
  for (var i = 0; i < uniq.length; i++) {{
    var o = document.createElement('option');
    o.value = uniq[i]; o.textContent = uniq[i];
    sel.appendChild(o);
  }}
}}
fillSelect('fGpa', allGpa);
fillSelect('fCat', allCat);
fillSelect('fYear', allYear);

var CYR2LAT_MAP = {{'а':'a','в':'b','е':'e','к':'k','м':'m','н':'h','о':'o','р':'p','с':'c','т':'t','у':'y','х':'x'}};
function canonicalizeTagRun(run) {{
  var out = '';
  for (var i=0;i<run.length;i++) {{ var ch = run[i]; out += CYR2LAT_MAP[ch] || ch; }}
  return out;
}}
function normalize(s) {{
  var t = (s || '').toLowerCase();
  t = t.replace(/\\b([a-zа-я]{{1,4}})[\\s-]+(\\d+)/gi,'$1$2');
  t = t.replace(/(\\d)[\\s-]+([a-zа-я]{{1,4}}\\d+)/gi,'$1$2');
  t = t.replace(/[«»"'.,()–—]/g, ' ');
  t = t.replace(/-/g,'');
  t = t.replace(/[a-zа-я0-9\\/]{{3,}}/gi, function(run){{
    if (/[0-9]/.test(run) && /[a-zа-я]/i.test(run)) {{ return canonicalizeTagRun(run); }}
    return run;
  }});
  return t.replace(/^\\s+|\\s+$/g, '');
}}
function tokenize(s) {{
  var norm = normalize(s);
  if (norm === '') return [];
  var parts = norm.split(/\\s+/);
  var out = [];
  for (var i = 0; i < parts.length; i++) {{ if (parts[i] !== '') out.push(parts[i]); }}
  return out;
}}
function isTagToken(t) {{ return t.length>=3 && /[0-9]/.test(t) && /[a-zа-я]/i.test(t); }}

var FIELD_WEIGHTS = [['cause', 4], ['name', 2], ['category', 1.5], ['act', 1], ['date', 1], ['gpaStr', 1], ['measures', 1], ['conclusion', 0.5], ['recommendation', 0.5]];
var TAG_BOOST = 9;

function scoreIncident(inc, tokens) {{
  if (tokens.length === 0) return 0.0001;
  var hasTags = false;
  for (var q = 0; q < tokens.length; q++) {{ if (isTagToken(tokens[q])) {{ hasTags = true; break; }} }}
  var fields = {{ name: inc.name, cause: inc.cause, category: inc.category, act: inc.act, date: inc.date, gpaStr: inc.gpa.join(' '), measures: inc.measures, conclusion: inc.conclusion, recommendation: inc.recommendation }};
  var score = 0;
  for (var t = 0; t < tokens.length; t++) {{
    var tok = tokens[t];
    var tag = isTagToken(tok);
    if (hasTags && !tag) continue;
    var boost = tag ? TAG_BOOST : 1;
    for (var f = 0; f < FIELD_WEIGHTS.length; f++) {{
      var fieldName = FIELD_WEIGHTS[f][0], weight = FIELD_WEIGHTS[f][1];
      var text = normalize(fields[fieldName]);
      if (text.indexOf(tok) !== -1) {{
        var mult = boost;
        if (fieldName==='cause') {{ mult *= (tag ? 2 : 1.5); }}
        score += weight*mult;
      }}
    }}
  }}
  return score;
}}

function highlight(text, tokens) {{
  var safe = escapeHtml(text);
  if (!tokens.length) return safe;
  for (var i = 0; i < tokens.length; i++) {{
    var tok = tokens[i];
    if (tok.length < 2) continue;
    var re = new RegExp('(' + escapeRegExp(tok) + ')', 'ig');
    safe = safe.replace(re, '<mark>$1</mark>');
  }}
  return safe;
}}

function render() {{
  var q = document.getElementById('q').value;
  var tokens = tokenize(q);
  var fGpa = document.getElementById('fGpa').value;
  var fCat = document.getElementById('fCat').value;
  var fYear = document.getElementById('fYear').value;

  var scored = [];
  for (var i = 0; i < incidents.length; i++) {{
    var inc = incidents[i];
    var score = scoreIncident(inc, tokens);
    if (score <= 0) continue;
    if (fGpa && !arrayContains(inc.gpa, fGpa)) continue;
    if (fCat && inc.category !== fCat) continue;
    if (fYear && inc.date.slice(-4) !== fYear) continue;
    scored.push({{ inc: inc, score: score }});
  }}
  scored.sort(function (a, b) {{ return b.score - a.score; }});

  document.getElementById('count').innerHTML = 'Найдено: <b>' + scored.length + '</b> из ' + incidents.length;
  var box = document.getElementById('results');
  box.innerHTML = '';

  if (scored.length === 0) {{
    box.innerHTML = '<div class="empty"><div class="big">&empty;</div>Ничего не найдено. Попробуйте другой запрос или сбросьте фильтры.</div>';
    return;
  }}

  for (var s = 0; s < scored.length; s++) {{
    var inc = scored[s].inc;
    var card = document.createElement('div');
    card.className = 'card';
    var gpaTags = '';
    for (var g = 0; g < inc.gpa.length; g++) {{ gpaTags += '<span class="tag gpa">' + escapeHtml(inc.gpa[g]) + '</span>'; }}
    var htmlStr = '';
    htmlStr += '<div class="card-top">';
    htmlStr += '<span class="tag act">' + escapeHtml(inc.act) + ' &middot; ' + escapeHtml(inc.date) + '</span>';
    htmlStr += gpaTags;
    htmlStr += '<span class="tag cat">' + highlight(inc.category, tokens) + '</span>';
    htmlStr += '<span class="tag type">' + escapeHtml(inc.type) + '</span>';
    htmlStr += '</div>';
    htmlStr += '<h3>' + highlight(inc.name, tokens) + '</h3>';
    htmlStr += '<p class="cause"><b>Обстоятельства, при которых произошел останов:</b> ' + highlight(inc.cause, tokens) + '</p>';
    htmlStr += '<button class="toggle-btn" type="button">Показать меры и заключение &#9662;</button>';
    htmlStr += '<div class="details">';
    htmlStr += '<p><b>Заключение:</b> ' + highlight(inc.conclusion, tokens) + '</p>';
    var principalMeasures = '';
    if (inc.remediation && inc.remediation.length) {{
      var pmParts = [];
      for (var w = 0; w < inc.remediation.length; w++) {{
        var wk = inc.remediation[w];
        pmParts.push((wk.work_description || '') + ' — Акт выполненных работ ' + (wk.work_act || '?') + ' от ' + (wk.work_date || '?') + ' (' + (wk.status || 'выполнено') + ')');
      }}
      principalMeasures = pmParts.join('<br>');
    }} else {{
      principalMeasures = inc.measures || '—';
    }}
    htmlStr += '<p><b>Принятые меры:</b> ' + highlight(principalMeasures, tokens) + '</p>';
    var docLinksAI = [];
    if (inc.source_uri) {{
      docLinksAI.push('<a class="doc-link" href="' + inc.source_uri + '" download="' + escapeHtml(inc.source_name || 'act.pdf') + '" target="_blank">Открыть акт расследования</a>');
    }}
    if (inc.remediation && inc.remediation.length) {{
      for (var w2 = 0; w2 < inc.remediation.length; w2++) {{
        var wk2 = inc.remediation[w2];
        var wkLink = wk2.work_source_uri || wk2.work_source || wk2.source;
        if (wkLink) {{
          docLinksAI.push('<a class="doc-link" href="' + wkLink + '" download="' + escapeHtml(wk2.work_source || 'work.pdf') + '" target="_blank">Открыть акт выполненных работ</a>');
        }}
      }}
    }}
    if (inc.defect_source) {{
      docLinksAI.push('<a class="doc-link" href="' + inc.defect_source + '" download="' + escapeHtml(inc.defect_source_name || 'defect.pdf') + '" target="_blank">Открыть дефектный акт</a>');
    }}
    if (inc.work_source) {{
      docLinksAI.push('<a class="doc-link" href="' + inc.work_source + '" download="' + escapeHtml(inc.work_source_name || 'work.pdf') + '" target="_blank">Открыть акт выполненных работ</a>');
    }}
    if (docLinksAI.length) {{
      htmlStr += '<p><b>Документы:</b> ' + docLinksAI.join(' &middot; ') + '</p>';
    }}
    htmlStr += '</div>';
    card.innerHTML = htmlStr;
    (function (cardEl) {{
      var btn = cardEl.querySelector('.toggle-btn');
      var det = cardEl.querySelector('.details');
      btn.addEventListener('click', function () {{
        var isOpen = det.className.indexOf('open') !== -1;
        if (isOpen) {{ det.className = 'details'; btn.innerHTML = 'Показать меры и заключение &#9662;'; }}
        else {{ det.className = 'details open'; btn.innerHTML = 'Скрыть детали &#9652;'; }}
      }});
    }})(card);
    box.appendChild(card);
  }}
}}

document.getElementById('q').addEventListener('input', render);
document.getElementById('fGpa').addEventListener('change', render);
document.getElementById('fCat').addEventListener('change', render);
document.getElementById('fYear').addEventListener('change', render);
document.getElementById('reset').addEventListener('click', function () {{
  document.getElementById('q').value = '';
  document.getElementById('fGpa').value = '';
  document.getElementById('fCat').value = '';
  document.getElementById('fYear').value = '';
  render();
}});
render();
</script>
</body>
</html>
"""


def build_html(incidents, defects):
    slim = []
    for inc in incidents:
        source_uri = pdf_to_data_uri(inc.get('source', '')) if inc.get('source') else None
        remediation_out = []
        for w in inc.get('remediation', []):
            w_out = dict(w)
            w_out['work_source_uri'] = pdf_to_data_uri(w.get('work_source', '')) if w.get('work_source') else None
            w_out['defect_source_uri'] = pdf_to_data_uri(w.get('defect_source', '')) if w.get('defect_source') else None
            remediation_out.append(w_out)
        slim.append({
            'kind': 'incident',
            'act': inc['act'], 'date': inc['date'], 'time': inc['time'],
            'gpa': inc['gpa'], 'type': inc['type'], 'category': inc['category'],
            'name': inc['name'], 'cause': inc['cause'], 'measures': inc['measures'],
            'conclusion': inc['conclusion'], 'recommendation': inc['recommendation'],
            'remediation': remediation_out,
            'source_uri': source_uri, 'source_name': inc.get('source', ''),
        })
    for d in defects:
        materials_txt = ''
        if d.get('materials'):
            materials_txt = ' Расходные материалы: ' + '; '.join(
                f"{m['name']} ({m['part_no']}) — {m['qty']} {m['unit']}" for m in d['materials'])
        defect_uri = pdf_to_data_uri(d.get('defect_source', '')) if d.get('defect_source') else None
        work_uri = pdf_to_data_uri(d.get('work_source', '')) if d.get('work_source') else None
        slim.append({
            'kind': 'defect',
            'act': d.get('defect_act', ''), 'date': d.get('defect_date', ''), 'time': '',
            'gpa': d.get('gpa', []), 'type': 'Дефект (без останова)',
            'category': 'Плановая инспекция / дефект',
            'name': d.get('title', d.get('description', '')[:80]),
            'cause': d.get('description', ''),
            'measures': '',
            'conclusion': 'Статус: ' + d.get('status', ''),
            'recommendation': d.get('work_description', '') + materials_txt,
            'remediation': [],
            'defect_source': defect_uri, 'defect_source_name': d.get('defect_source', ''),
            'work_source': work_uri, 'work_source_name': d.get('work_source', ''),
            'work_act': d.get('work_act', ''), 'work_date': d.get('work_date', ''),
        })
    incidents_json = json.dumps(slim, ensure_ascii=False, indent=0)
    html_content = HTML_TEMPLATE.format(incidents_json=incidents_json, n_records=len(slim))
    with open(HTML_OUT, 'w', encoding='utf-8') as f:
        f.write(html_content)


if __name__ == '__main__':
    incidents = load_data()
    defects = load_defects()
    last_row = build_xlsx(incidents, defects)
    build_html(incidents, defects)
    build_web_site(incidents, defects, BASE_DIR / 'site')
    print(f'Готово: {len(incidents)} аварий, {len(defects)} отдельных дефектов -> '
          f'{XLSX_OUT.name}, {HTML_OUT.name}, site/index.html')
